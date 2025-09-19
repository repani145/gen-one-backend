# Python standard libraries
import os
import io
import glob
import shutil
import tempfile
import threading
import time
import uuid
from datetime import datetime
from collections import defaultdict

# Third-party libraries
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import django_filters

# Django core
from django.conf import settings
from django.db import IntegrityError,transaction
from django.db.models import Q, Max, Exists, OuterRef
from django.core.exceptions import ObjectDoesNotExist
from django.core.mail import EmailMessage
from django.core.signing import TimestampSigner, BadSignature, SignatureExpired
from django.http import (
    HttpResponse,
    JsonResponse,
    FileResponse,
    HttpResponseForbidden,
    Http404,
)
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

# Django REST framework
from rest_framework import status, generics, viewsets, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework_simplejwt.authentication import JWTAuthentication
from .pagination import StandardResultsSetPagination
from rest_framework.exceptions import PermissionDenied


# Local imports
from . import models, serializers, messages, validators, constants
from .exceptions import SerializerError
from .models import DataObject, DataFile, Specs, ApprovalComment
from .serializers import (
    FileUploadSerializer,
    DataObjectSerializer,
    SpecsSerializer,
    DataFileSerializer,
)
from .permissions import DataObjectWriteLockPermission
from .file_utils import (
    get_file_path_with_object_name,
    get_file_paths,
    get_target_specs,
)
from .working_file_manager import (
    create_and_get_working_file_path,
    delete_working_directory,
)
from .CustomValidationFiles.common_rules_validators import run_default_validators
from .CustomValidationFiles.custom_rule_validator import run_custom_rule_validation

import tempfile
from django.core.files import File
import uuid, threading, time
from django.core.files.base import File
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import pandas as pd, tempfile
from .models import DataObject, Specs

import pandas as pd
import openpyxl
import io
from django.core.files.uploadedfile import InMemoryUploadedFile
from sqlalchemy import inspect
import pymysql
from math import isfinite


import os
from decouple import config

SAP_DB_USER = config("SAP_DB_USER")
SAP_DB_PASS = config("SAP_DB_PASSWORD")
SAP_DB_HOST = config("SAP_DB_HOST")
SAP_DB_PORT = config("SAP_DB_PORT")
SAP_SAP_DB_NAME = config("SAP_DB_NAME")



class DataObjectViewSet(viewsets.ModelViewSet):
    queryset = DataObject.objects.all()
    serializer_class = DataObjectSerializer
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

class SpecsViewSet(viewsets.ModelViewSet):
    queryset = Specs.objects.all()
    serializer_class = SpecsSerializer
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

class DataFileViewSet(viewsets.ModelViewSet):
    queryset = DataFile.objects.all()
    serializer_class = DataFileSerializer
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

class RuleAppliedFilter(django_filters.FilterSet):
    # create a custom alias for the long FK chain
    objectName = django_filters.CharFilter(
        field_name="spec__objectName__objectName", lookup_expr="exact"
    )

     # ✅ New filter for spec_id
    spec_id = django_filters.NumberFilter(
        field_name="spec__id", lookup_expr="exact"
    )

    created_after = django_filters.DateFilter(
        field_name="created_at", lookup_expr="gte"
    )
    created_before = django_filters.DateFilter(
        field_name="created_at", lookup_expr="lte"
    )

    class Meta:
        model = models.RuleApplied
        fields = ["objectName", "rule_applied", "created_after", "created_before"]

class RuleAppliedGetView(APIView):
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, *args, **kwargs):
        queryset = models.RuleApplied.objects.select_related("spec").all()

        # Apply filters manually
        filterset = RuleAppliedFilter(request.GET, queryset=queryset)
        if not filterset.is_valid():
            return Response(filterset.errors, status=status.HTTP_400_BAD_REQUEST)
        queryset = filterset.qs


        # Apply ordering
        ordering = request.GET.get("ordering")
        if ordering:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by("-created_at")  # default ordering

        # ✅ Apply pagination
        paginator = StandardResultsSetPagination()
        paginated_qs = paginator.paginate_queryset(queryset, request, view=self)

        serializer = serializers.RuleAppliedTableSerializer(paginated_qs, many=True)
        return paginator.get_paginated_response(serializer.data)

class AllDataObjectView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request):
        context = {
            "success": 1,
            "message": messages.DATA_FOUND,
            "data": {}
        }
        try:
            modules = models.DataObject.objects.all()
            module_ser = serializers.DataObjectSerializer(modules, many=True)
            context['data'] = module_ser.data
        except Exception as e:
            context['success'] = 0
            context['message'] = str(e)
        return Response(context)

class AllDependenciesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request):
        context = {
            "success": 1,
            "message": messages.DATA_FOUND,
            "data": []
        }
        try:
            # Collect all objectName values
            dependencies = list(models.DataObject.objects.values_list("objectName", flat=True))

            # Remove duplicates & sort
            unique_dependencies = sorted(set(dependencies))

            context["data"] = unique_dependencies

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

class DataObjectView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

    def get(self, request, id):
        context = {"success": 1, "message": messages.DATA_FOUND, "data": []}
        try:
            print("➡️ GET request for DataObject ID:", id)

            module = models.DataObject.objects.get(id=id)

            # ✅ Permission check
            self.check_object_permissions(request, module)

            module_ser = serializers.DataObjectSerializer(module)
            context["data"] = module_ser.data
            context["count"] = 1

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def post(self, request, *args, **kwargs):
        context = {"success": 1, "message": messages.DATA_SAVED, "data": {}}
        try:
            print("➡️ POST request with data:", request.data)

            validator = validators.ObjectDataValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            req_params = validator.validated_data
            module_instance = models.DataObject(**req_params)

            # # ✅ Permission check (new object)
            # self.check_object_permissions(request, module_instance)

            module_instance.clean()
            module_instance.save()

            context["data"] = serializers.DataObjectSerializer(module_instance).data

        except IntegrityError as e:
            context["success"] = 0
            if "Duplicate entry" in str(e):
                err_msg = str(e)
                duplicate_value = err_msg.split("'")[1]
                field_name = err_msg.split("for key")[1].split("'")[1]
                context["message"] = f"❌ '{duplicate_value}' already exists."
            else:
                context["message"] = "❌ Database Integrity Error."
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def put(self, request, id):
        context = {"success": 1, "message": messages.DATA_UPDATED, "data": {}}
        try:
            print("➡️ PUT request for ID:", id, "Data:", request.data)

            if not id:
                raise Exception("ID is required for update")

            validator = validators.ObjectDataValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            module_obj = models.DataObject.objects.get(id=id)

            # ✅ Permission check
            self.check_object_permissions(request, module_obj)

            for field in ["company", "dependencies", "module", "objectName"]:
                if field in request.data:
                    setattr(module_obj, field, request.data[field])

            module_obj.save()
            context["data"] = serializers.DataObjectSerializer(module_obj).data

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def patch(self, request, *args, **kwargs):
        context = {"success": 1, "message": messages.DATA_UPDATED, "data": {}}
        try:
            record_id = request.data.get("id")
            print("➡️ PATCH request for ID:", record_id, "Data:", request.data)

            if not record_id:
                raise Exception("ID is required for partial update")

            module_obj = models.DataObject.objects.get(id=record_id)

            # ✅ Permission check
            self.check_object_permissions(request, module_obj)

            for field in ["company", "dependencies", "module", "objectName"]:
                if field in request.data:
                    setattr(module_obj, field, request.data[field])

            module_obj.full_clean()
            module_obj.save()
            context["data"] = serializers.DataObjectSerializer(module_obj).data

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def delete(self, request, id):
        context = {"success": 1, "message": messages.DATA_DELETED, "data": {}}
        try:
            print("➡️ DELETE request for ID:", id)

            if not id:
                raise Exception("ID is required for delete")

            module_obj = models.DataObject.objects.get(id=id)

            # ✅ Permission check
            self.check_object_permissions(request, module_obj)

            # 🔍 Check if this object is a dependency in other objects
            dependent_objects = models.DataObject.objects.filter(
                dependencies__contains=[module_obj.objectName]
            )
            if dependent_objects.exists():
                raise Exception(
                    f"Cannot delete '{module_obj.objectName}' because it is a dependency of other objects."
                )

            module_obj.delete()

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

# new
class SpecsAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

    def get(self, request, *args, **kwargs):
        context = {"success": 1, "message": messages.DATA_FOUND, "data": {}}
        try:
            object_id = kwargs.get("id")
            if not object_id:
                context["success"] = 0
                context["message"] = "objectName_id is required"
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

            # 🔹 Fetch object and check permission
            data_object = get_object_or_404(models.DataObject, id=object_id)
            self.check_object_permissions(request, data_object)

            specs = (
                models.Specs.objects
                .select_related("objectName")
                .filter(objectName_id=object_id)
                .order_by("tab", "position")
            )

            if not specs.exists():
                context["success"] = 1
                context["message"] = "No specs data exists"
                context["data"] = {"objectName": data_object.objectName}
                return Response(context, status=status.HTTP_200_OK)

            grouped_data = defaultdict(list)
            for spec in specs:
                rules_qs = models.RuleApplied.objects.filter(spec=spec)
                rules_data = [
                    {
                        "id": rule.id,
                        "rule_applied": rule.rule_applied,
                        "description": rule.description,
                        "rule_applied_data": rule.rule_applied_data,
                    }
                    for rule in rules_qs
                ]

                grouped_data[spec.tab].append({
                    "id": spec.id,
                    "company": spec.company,
                    "field_id": (spec.field_id).upper(),
                    "mandatory": spec.mandatory,
                    "allowed_values": spec.allowed_values,
                    "sap_table": spec.sap_table,
                    "sap_field_id": spec.sap_field_id,
                    "sap_description": spec.sap_description,
                    "position": spec.position,
                    "rules": rules_data,
                })

            tab_data = [{"tab": tab_name.lower(), "fields": fields} for tab_name, fields in grouped_data.items()]

            response_data = {
                "objectName": specs.first().objectName.objectName,
                "tabs": tab_data,
            }

            context["success"] = 1
            context["message"] = "Data fetched successfully"
            context["data"] = response_data
            return Response(context, status=status.HTTP_200_OK)

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            context["data"] = {}
            return Response(context, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request, *args, **kwargs):
        context = {"success": 1, "message": "Data saved successfully", "data": {}}
        try:
            validator = validators.SpecsValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            req_params = validator.validated_data
            data_object = get_object_or_404(models.DataObject, pk=req_params["objectName"])

            # 🔹 Check write permission
            self.check_object_permissions(request, data_object)

            req_params["objectName"] = data_object

            # Auto-assign position based on tab
            existing_fields = models.Specs.objects.filter(
                objectName=data_object, tab=req_params["tab"]
            ).order_by("position")

            req_params["position"] = existing_fields.last().position + 1 if existing_fields.exists() else 0

            spec_instance = models.Specs.objects.create(**req_params)
            context["data"] = serializers.SpecsSerializer(spec_instance).data

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
        return Response(context)

    def put(self, request, id=None, *args, **kwargs):
        context = {"success": 1, "message": "Data updated successfully", "data": {}}
        try:
            spec_instance = get_object_or_404(models.Specs, pk=id)

            # 🔹 Check write permission
            self.check_object_permissions(request, spec_instance)

            validator = validators.SpecsUpdateValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            req_params = validator.validated_data

            if "objectName" in req_params:
                data_object = get_object_or_404(models.DataObject, pk=req_params["objectName"])
                self.check_object_permissions(request, data_object)
                req_params["objectName"] = data_object

            if "allowed_values" in req_params and isinstance(req_params["allowed_values"], str):
                req_params["allowed_values"] = [
                    v.strip() for v in req_params["allowed_values"].split(",") if v.strip()
                ]

            for key, value in req_params.items():
                setattr(spec_instance, key, value)
            spec_instance.save()

            context["data"] = serializers.SpecsSerializer(spec_instance).data

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
        return Response(context)

    def delete(self, request, id=None):
        context = {"success": 1, "message": messages.DATA_DELETED, "data": {}}
        try:
            spec_instance = get_object_or_404(models.Specs, pk=id)

            # 🔹 Check write permission
            self.check_object_permissions(request, spec_instance)

            spec_instance.delete()

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
        return Response(context)

# neww added release or request processing permission check
class ReorderSpecsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

    def put(self, request, *args, **kwargs):
        try:
            object_id = request.data.get("objectName_id")
            tab = request.data.get("tab")
            fields = request.data.get("fields", [])

            if not object_id or not tab or not fields:
                return Response({"success": 0, "message": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)

            # 🔹 Fetch DataObject and check permission
            data_object = get_object_or_404(models.DataObject, id=object_id)
            self.check_object_permissions(request, data_object)

            # update positions
            for field in fields:
                field_id = field.get("id")
                position = field.get("position")
                if field_id is not None and position is not None:
                    models.Specs.objects.filter(
                        id=field_id,
                        objectName_id=object_id,
                        tab=tab
                    ).update(position=position)

            return Response({"success": 1, "message": "Reordering saved successfully"}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"success": 0, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# OLD
# class ReorderSpecsView(APIView):
#     authentication_classes = [JWTAuthentication]
#     permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

#     def put(self, request, *args, **kwargs):
#         try:
#             object_id = request.data.get("objectName_id")
#             tab = request.data.get("tab")
#             fields = request.data.get("fields", [])

#             if not object_id or not tab or not fields:
#                 return Response({"success": 0, "message": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)

#             # update positions
#             for field in fields:
#                 field_id = field.get("id")
#                 position = field.get("position")
#                 if field_id is not None and position is not None:
#                     models.Specs.objects.filter(
#                         id=field_id,
#                         objectName_id=object_id,
#                         tab=tab
#                     ).update(position=position)

#             return Response({"success": 1, "message": "Reordering saved successfully"}, status=status.HTTP_200_OK)

#         except Exception as e:
#             return Response({"success": 0, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# not using
class SpecsTemplateDownloadAPIView(APIView):
    # authentication_classes = [JWTAuthentication]
    # permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, *args, **kwargs):
        context = {"success": 1, "message": "Template generated successfully", "data": {}}
        try:
            # ✅ Create workbook & sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Specs Template"

            # ✅ Define Specs model fields (exclude id/auto fields if required)
            fields = [
                "company",
                "objectName",
                "tab",
                "field_id",   
                "mandatory",        # Yes / No
                "allowed_values",   # Comma separated            
                "sap_table",
                "sap_field_id",
                "sap_description",
            ]

            # ✅ Write header row
            ws.append(fields)

            # ✅ Add dropdown validation for "mandatory" column
            # Column H = 8th column (mandatory field)
            dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            ws.add_data_validation(dv)

            # Apply dropdown for a reasonable number of rows (say 1000)
            dv.add(f"E2:E1000")

            # (Optional) Example row for user guidance
            # ws.append([
            #     "f101", "38", "GenOne", "tab1",
            #     "t11", "s11", "desc...", "Yes",
            #     "11,22,33", "1",
            # ])

            # ✅ Prepare Excel response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="SpecsTemplate.xlsx"'
            wb.save(response)

            return response

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            return JsonResponse(context, status=500)


# not using
class SpecsTemplateUploadAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def post(self, request, *args, **kwargs):
        context = {"success": 1, "message": "Data uploaded successfully", "data": []}
        try:
            file_obj = request.FILES.get("file")
            if not file_obj:
                raise Exception("No file uploaded")

            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active

            # ✅ Expected header row
            expected_headers = [
                "field_id",
                "objectName_id",
                "company",
                "tab",
                "sap_table",
                "sap_field_id",
                "sap_description",
                "mandatory",
                "allowed_values",
                "position",
            ]

            headers = [cell.value for cell in ws[1]]
            if headers != expected_headers:
                raise Exception("Invalid template format. Please use the downloaded template.")

            # ✅ Iterate rows and save
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # skip empty rows
                    continue

                (
                    field_id,
                    objectName_id,
                    company,
                    tab,
                    sap_table,
                    sap_field_id,
                    sap_description,
                    mandatory,
                    allowed_values,
                    position,
                ) = row

                # ✅ handle mandatory Yes/No → Boolean
                mandatory_val = True if str(mandatory).lower() == "yes" else False

                # ✅ handle allowed_values → list
                allowed_list = (
                    [v.strip() for v in str(allowed_values).split(",") if v.strip()]
                    if allowed_values
                    else None
                )

                # ✅ ensure DataObject exists
                try:
                    data_object = models.DataObject.objects.get(objectName=objectName_id)
                except ObjectDoesNotExist:
                    data_object = models.DataObject.objects.create(objectName=objectName_id)

                # ✅ create or update Specs
                spec_obj, created = models.Specs.objects.update_or_create(
                    field_id=field_id,
                    defaults={
                        "objectName": data_object,
                        "company": company,
                        "tab": tab,
                        "sap_table": sap_table,
                        "sap_field_id": sap_field_id,
                        "sap_description": sap_description,
                        "mandatory": mandatory_val,
                        "allowed_values": allowed_list,
                        "position": position,
                    },
                )

                context["data"].append(serializers.SpecsSerializer(spec_obj).data)

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return JsonResponse(context)


class CustomRuleTemplateUIViewSet(viewsets.ModelViewSet):
    """
    Provides: list, retrieve (by pk), create, update, partial_update, destroy
    Extra action: retrieve_by_rule (GET /api/rules/schema/<rule_name>/)
    """

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    queryset = models.CustomRuleTemplateUI.objects.all()
    serializer_class = serializers.CustomRuleTemplateUISerializer
    lookup_field = "pk"

    @action(detail=False, methods=["get"], url_path=r"schema/(?P<rule_name>[^/.]+)")
    def retrieve_by_rule(self, request, rule_name=None):
        """
        GET /api/rules/schema/<rule_name>/
        returns the schema JSON for the given rule_name
        """
        obj = get_object_or_404(models.CustomRuleTemplateUI, rule_name=rule_name)
        serializer = self.get_serializer(obj)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AllRulesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request):
        context = {
            "success": 1,
            "message": "Rule names fetched successfully",
            "data": []
        }
        try:
            rules = models.CustomRuleTemplateUI.objects.all()
            rule_names = rules.values_list("rule_name", flat=True)  # gives QuerySet
            context["data"] = list(rule_names)  # convert to list for JSON response
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)


# old
# class RuleAppliedView(APIView):
#     authentication_classes = [JWTAuthentication]
#     permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

#     # CREATE (POST)
#     def post(self, request):
#         context = {
#             "success": 1,
#             "message": "Rule saved successfully",
#             "data": {}
#         }
#         try:
#             data = request.data
#             # 1️⃣ Extract source_fields from payload
#             source_fields = data.get("rule_applied_data", {}).get("source", {})
#             spec_id = models.DataObject.objects.get(objectName=source_fields.get("spec")).id
#             print('sourec fields.............................\n',spec_id)

#             # 2️⃣ Find matching Specs record
#             spec_obj = models.Specs.objects.filter(
#                 objectName=spec_id,
#                 tab=source_fields.get("tab"),
#                 field_id=source_fields.get("field"),
#             ).first()

#             if not spec_obj:
#                 context["success"] = 0
#                 context["message"] = "Spec not found for given source_fields"
#                 return Response(context, status=status.HTTP_400_BAD_REQUEST)

#             # 3️⃣ Inject spec_field_id before serializer validation
#             data["spec"] = spec_obj.id

#             serializer = serializers.RuleAppliedSerializer(
#                 data=data,
#                 context={"request": request}
#             )
#             if serializer.is_valid():
#                 serializer.save()
#                 context["data"] = serializer.data
#             else:
#                 context["success"] = 0
#                 context["message"] = "Validation failed"
#                 context["errors"] = serializer.errors
#                 return Response(context, status=status.HTTP_400_BAD_REQUEST)

#         except Exception as e:
#             context["success"] = 0
#             context["message"] = str(e)
#             return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         return Response(context, status=status.HTTP_201_CREATED)

#     # READ (GET Single Rule by ID)
#     def get(self, request, pk=None):
#         try:
#             rule = models.RuleApplied.objects.select_related("spec").get(id=pk)
#             serializer = serializers.RuleAppliedSerializer(rule)
#             return Response(serializer.data, status=status.HTTP_200_OK)
#         except models.RuleApplied.DoesNotExist:
#             return Response(
#                 {"success": 0, "message": "Rule not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

#     # UPDATE (PUT/PATCH)
#     def put(self, request, pk=None):
#         context = {"success": 1, "message": "Rule updated successfully", "data": {}}
#         try:
#             rule = models.RuleApplied.objects.get(id=pk)
#             serializer = serializers.RuleAppliedSerializer(
#                 rule, data=request.data, partial=True, context={"request": request}
#             )
#             if serializer.is_valid():
#                 serializer.save()
#                 context["data"] = serializer.data
#             else:
#                 context["success"] = 0
#                 context["message"] = "Validation failed"
#                 context["errors"] = serializer.errors
#                 return Response(context, status=status.HTTP_400_BAD_REQUEST)

#         except models.RuleApplied.DoesNotExist:
#             return Response(
#                 {"success": 0, "message": "Rule not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             context["success"] = 0
#             context["message"] = str(e)
#             return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         return Response(context, status=status.HTTP_200_OK)

#     # DELETE
#     def delete(self, request, pk=None):
#         try:
#             rule = models.RuleApplied.objects.get(id=pk)
#             rule.delete()
#             return Response(
#                 {"success": 1, "message": "Rule deleted successfully"},
#                 status=status.HTTP_200_OK,
#             )
#         except models.RuleApplied.DoesNotExist:
#             return Response(
#                 {"success": 0, "message": "Rule not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"success": 0, "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )

# new
class RuleAppliedView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

    # CREATE (POST)
    def post(self, request):
        context = {"success": 1, "message": "Rule saved successfully", "data": {}}
        try:
            data = request.data
            # Extract source_fields from payload
            source_fields = data.get("rule_applied_data", {}).get("source", {})
            data_object_id = models.DataObject.objects.get(objectName=source_fields.get("spec")).id

            # Find matching Specs record
            spec_obj = models.Specs.objects.filter(
                objectName=data_object_id,
                tab=source_fields.get("tab"),
                field_id=source_fields.get("field"),
            ).first()

            if not spec_obj:
                context["success"] = 0
                context["message"] = "Spec not found for given source_fields"
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

            # 🔹 Check write permission on the related DataObject
            self.check_object_permissions(request, spec_obj)

            # Inject spec_field_id before serializer validation
            data["spec"] = spec_obj.id
            serializer = serializers.RuleAppliedSerializer(data=data, context={"request": request})

            if serializer.is_valid():
                serializer.save()
                context["data"] = serializer.data
            else:
                context["success"] = 0
                context["message"] = "Validation failed"
                context["errors"] = serializer.errors
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(context, status=status.HTTP_201_CREATED)

    # READ (GET Single Rule by ID)
    def get(self, request, pk=None):
        try:
            rule = models.RuleApplied.objects.select_related("spec").get(id=pk)
            serializer = serializers.RuleAppliedSerializer(rule)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except models.RuleApplied.DoesNotExist:
            return Response({"success": 0, "message": "Rule not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # UPDATE (PUT/PATCH)
    def put(self, request, pk=None):
        context = {"success": 1, "message": "Rule updated successfully", "data": {}}
        try:
            rule = get_object_or_404(models.RuleApplied, pk=pk)

            # 🔹 Check write permission via related Spec → DataObject
            self.check_object_permissions(request, rule)

            serializer = serializers.RuleAppliedSerializer(
                rule, data=request.data, partial=True, context={"request": request}
            )
            if serializer.is_valid():
                serializer.save()
                context["data"] = serializer.data
            else:
                context["success"] = 0
                context["message"] = "Validation failed"
                context["errors"] = serializer.errors
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

        except models.RuleApplied.DoesNotExist:
            return Response({"success": 0, "message": "Rule not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(context, status=status.HTTP_200_OK)

    # DELETE
    def delete(self, request, pk=None):
        try:
            rule = get_object_or_404(models.RuleApplied, pk=pk)

            # 🔹 Check write permission via related Spec → DataObject
            self.check_object_permissions(request, rule)

            rule.delete()
            return Response({"success": 1, "message": "Rule deleted successfully"}, status=status.HTTP_200_OK)

        except models.RuleApplied.DoesNotExist:
            return Response({"success": 0, "message": "Rule not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({"success": 0, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RuleAppliedListView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, *args, **kwargs):
        try:
            queryset = models.RuleApplied.objects.select_related("spec").all()

            # --- Filtering ---
            object_name = request.GET.get("objectName")
            if object_name:
                queryset = queryset.filter(spec__objectName__objectName=object_name)

            rule_applied = request.GET.get("rule_applied")
            if rule_applied:
                queryset = queryset.filter(rule_applied__icontains=rule_applied)

            created_after = request.GET.get("created_after")
            if created_after:
                queryset = queryset.filter(created_at__gte=created_after)

            created_before = request.GET.get("created_before")
            if created_before:
                queryset = queryset.filter(created_at__lte=created_before)

            # --- Search ---
            search = request.GET.get("search")
            if search:
                queryset = queryset.filter(
                    Q(description__icontains=search)
                    | Q(spec__tab__icontains=search)
                    | Q(spec__field_id__icontains=search)
                )

            # --- Ordering ---
            ordering = request.GET.get("ordering")
            if ordering:
                queryset = queryset.order_by(ordering)
            else:
                queryset = queryset.order_by("-created_at")

            serializer = serializers.RuleAppliedTableSerializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class RuleAppliedBySpecView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, spec_id):
        context = {
            "success": 1,
            "message": "Rules fetched successfully",
            "data": []
        }
        try:
            # SQLite-safe: values() + distinct()
            queryset = (
                models.RuleApplied.objects.filter(spec_id=spec_id)
                .values("rule_applied")
                .distinct()
            )

            serializer = serializers.RuleAppliedNameSerializer(queryset, many=True)

            # Extract only values → ["rule1", "rule2", ...]
            context["data"] = [item["rule_applied"] for item in serializer.data]

        except Exception as e:
            context["success"] = 0
            context["message"] = f"Failed to fetch rules: {str(e)}"

        return Response(context, status=status.HTTP_200_OK)

def clean_excel(uploaded_file):
    """Return a cleaned Excel file object without 'mapping' sheet."""
    xls = pd.ExcelFile(uploaded_file)

    # write cleaned file to memory (BytesIO)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name in xls.sheet_names:
            if sheet_name.lower() != "mapping":
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)  # reset pointer to start

    # Wrap as InMemoryUploadedFile so Django treats it like a real upload
    cleaned_file = InMemoryUploadedFile(
        file=output,
        field_name="file",
        name="cleaned_file.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size=output.getbuffer().nbytes,
        charset=None,
    )

    return cleaned_file


def handle_validated_upload(uploaded_file, obj, object_name):
    # ✅ Ensure directory exists
    base_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower())
    archive_dir = os.path.join(base_dir, "archive")
    os.makedirs(base_dir, exist_ok=True)
    os.makedirs(archive_dir, exist_ok=True)


    # print(os.getcwd())
    print(os.listdir(base_dir))

    # Paths
    main_file_path = os.path.join(base_dir, f"{object_name}.xlsx")

    # 🔥 Check if a main file already exists
    if os.path.exists(main_file_path):
        # Get last version
        last_file = models.DataFile.objects.filter(data_object=obj).order_by("-version").first()
        next_version = last_file.version + 1 if last_file else 1

        # Move existing main file → archive
        archive_name = f"{object_name}_v{next_version}.xlsx"
        archive_path = os.path.join(archive_dir, archive_name)
        shutil.move(main_file_path, archive_path)

        # Update previous DB record version
        models.DataFile.objects.filter(
            data_object=obj, file_name=f"{object_name}.xlsx"
        ).update(file_name=archive_name, version=next_version,status='UPLOAD SUCCESS & READY FOR VALIDATION')

    # ✅ Save the new file as main (objectName.xlsx)
    with open(main_file_path, "wb+") as dest:
        for chunk in uploaded_file.chunks():
            dest.write(chunk)

    # 🔥 Create DB record for the new main file (no version)
    models.DataFile.objects.create(
        data_object=obj,
        file_name=f"{object_name}.xlsx",
        status=constants.STATUS_UPLOAD_SUCCESS,
        version=0,  # version 0 means "main/latest"
    )

    return Response(
        {
            "success": 1,
            "message": f"File uploaded and validated successfully for {object_name}",
            "file_name": f"{object_name}.xlsx",
            "status": constants.STATUS_UPLOAD_SUCCESS,
        },
        status=status.HTTP_200_OK,
    )

# In-memory progress store (better: Redis/DB in production)
UPLOAD_PROGRESS = {}
# OLD
class FileUploadAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, id=None):
        """
        If object_id is given → return files for that DataObject.
        Otherwise → return all uploaded files.
        """
        context = {
            "success": 1,
            "message": "",
            "data": [],
        }

        try:
            object_id = id
            if object_id:
                files = DataFile.objects.filter(data_object_id=object_id).order_by("-uploaded_at")
                context["message"] = f"Files for DataObject ID {object_id} retrieved successfully"
            else:
                files = DataFile.objects.filter(version=0).order_by("-uploaded_at")
                context["message"] = "All files retrieved successfully"

            print('files...->>>>>>>>>>>>>>>>>>>>\n',files)
            serializer = serializers.DataFileSerializer(files, many=True)
            context["data"] = serializer.data

            return Response(context, status=status.HTTP_200_OK)

        except Exception as e:
            context["success"] = 0
            context["message"] = f"Error: {str(e)}"
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        # print("📥 Received POST request to handle-file API")
        serializer = FileUploadSerializer(data=request.data)
        if not serializer.is_valid():
            # print("❌ Serializer validation failed:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        


        object_name = serializer.validated_data["objectName"]
        uploaded_file = serializer.validated_data["file"]
        print(f"✅ Serializer validated | objectName={object_name}, file={uploaded_file.name}")

        # 🔹 Get related DataObject
        data_object = models.DataObject.objects.filter(objectName=object_name).first()
        if not data_object:
            return Response(
                {"success": 0, "message": f"DataObject '{object_name}' not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        # 🔹 Run object-level permission check
        try:
            self.check_object_permissions(request, data_object)
        except PermissionDenied as e:
            return Response(
                {"success": 0, "message": str(e)},
                status=status.HTTP_403_FORBIDDEN
    )


        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_file_path = tmp.name
            for chunk in uploaded_file.chunks():
                tmp.write(chunk)
        # print(f"📂 Temp file saved at {temp_file_path}")

        # Validate objectName exists
        try:
            obj = models.DataObject.objects.get(objectName=object_name)
            # print(f"✅ DataObject found: {obj}")
        except models.DataObject.DoesNotExist:
            # print(f"❌ DataObject '{object_name}' not found in DB")
            return Response(
                {"success": 0, "message": f"Object '{object_name}' not found"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Specs check
        specs = Specs.objects.filter(objectName=obj)
        if not specs.exists():
            # print(f"❌ No specs defined for '{object_name}'")
            return Response(
                {"success": 0, "message": f"No specs defined for '{object_name}'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # print(f"✅ Specs found for '{object_name}' | count={specs.count()}")

        # Create unique upload ID
        upload_id = str(uuid.uuid4())
        UPLOAD_PROGRESS[upload_id] = {"progress": 0, "message": "Started processing...", "success": None}
        # print(f"🚀 Upload started | upload_id={upload_id}")

        # Process file in background
        def process_file():
            try:
                # Step 1: Building expected tabs
                # print(f"[{upload_id}] 🔄 Step 1: Building expected tabs")
                for p in range(1, 11):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Validating specs...", "success": None}
                    time.sleep(0.05)

                expected_tabs = {}
                for s in specs:
                    expected_tabs.setdefault(s.tab.lower(), set()).add(s.field_id.upper())
                # print(f"[{upload_id}] ✅ Expected tabs: {list(expected_tabs.keys())}")

                # Step 2: Reading Excel
                # print(f"[{upload_id}] 🔄 Reading Excel file {temp_file_path}")
                try:
                    xls = pd.ExcelFile(temp_file_path)
                    # print(f"[{upload_id}] ✅ Sheets found: {xls.sheet_names}")
                except Exception as e:
                    # print(f"[{upload_id}] ❌ Failed to read Excel file:", e)
                    UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": "Failed to read Excel file", "success": 0}
                    return

                for p in range(11, 31):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Preparing file...", "success": None}
                    time.sleep(0.05)

                # Step 3: Normalizing
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    cleaned_file_path = tmp.name
                    with pd.ExcelWriter(cleaned_file_path, engine="openpyxl") as writer:
                        for sheet_name in xls.sheet_names:
                            if sheet_name.lower() == "mapping":
                                print(f"[{upload_id}] ⚠️ Skipping 'mapping' sheet")
                                continue
                            df = pd.read_excel(xls, sheet_name=sheet_name)
                            df.columns = [str(col).upper() for col in df.columns]
                            df.to_excel(writer, sheet_name=sheet_name.lower(), index=False)
                    # print(f"[{upload_id}] ✅ Normalized file written to {cleaned_file_path}")

                for p in range(31, 61):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Validating tabs & fields...", "success": None}
                    time.sleep(0.05)

                # Step 4: Validate tabs
                cleaned_xls = pd.ExcelFile(cleaned_file_path)
                file_tabs = set(cleaned_xls.sheet_names)
                # print(f"[{upload_id}] ✅ Normalized file sheets: {file_tabs}")

                expected_tab_names = set(expected_tabs.keys())
                missing_tabs = expected_tab_names - file_tabs
                if missing_tabs:
                    # print(f"[{upload_id}] ❌ Missing tabs: {missing_tabs}")
                    UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": f"Missing required tabs...❌", "success": 0}
                    return

                for tab, expected_fields in expected_tabs.items():
                    df = pd.read_excel(cleaned_xls, sheet_name=tab, nrows=1)
                    file_fields = set(df.columns.astype(str).str.upper())
                    missing_fields = expected_fields - file_fields
                    if missing_fields:
                        # print(f"[{upload_id}] ❌ Missing fields in {tab}: {missing_fields}")
                        UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": f"Missing fields in '{tab}'", "success": 0}
                        return
                # print(f"[{upload_id}] ✅ All required fields present")

                for p in range(61, 91):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Saving validated file...", "success": None}
                    time.sleep(0.05)

                # Step 5: Save normalized file
                with open(cleaned_file_path, "rb") as f:
                    cleaned_file = File(f, name=f"{object_name.lower()}_cleaned.xlsx")
                    handle_validated_upload(cleaned_file, obj, object_name)
                # print(f"[{upload_id}] 📤 File uploaded successfully to DB/storage")

                # Step 6: Finalizing
                for p in range(91, 101):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Finalizing...", "success": None}
                    time.sleep(0.05)

                UPLOAD_PROGRESS[upload_id] = {"progress": 100, "message": "Upload completed successfully! 🎉", "success": 1}
                # print(f"[{upload_id}] 🎉 Upload finished successfully")

            except Exception as e:
                # print(f"[{upload_id}] ❌ Exception during processing:", e)
                UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": str(e), "success": 0}

        threading.Thread(target=process_file).start()
        return Response({"success": 1, "upload_id": upload_id}, status=status.HTTP_200_OK)

    def delete(self, request, id):
        try:
            file_obj = DataFile.objects.get(pk=id)

            try:
                self.check_object_permissions(request, file_obj)
            except PermissionDenied as e:
                return Response(
                    {"success": 0, "message": str(e)},
                    status=status.HTTP_403_FORBIDDEN
                )

            temp = ''
            for i in str(file_obj.file_name):
                if i!='.':
                    temp = temp + i
                else:
                    break
            file_name = str(file_obj.file_name)   # e.g., "employees.xlsx"
            media_root = settings.MEDIA_ROOT + str(temp)     # e.g., "FileData/customer"
            file_path = os.path.join(media_root, file_name)  # Full path with filename

            archive_dir = os.path.join(media_root, "archive")
            deleted_dir = os.path.join(archive_dir, "delete")

            os.makedirs(archive_dir, exist_ok=True)
            os.makedirs(deleted_dir, exist_ok=True)

            # 🔹 Find last version
            last_version = (
                DataFile.objects.filter(data_object=file_obj.data_object)
                .order_by("-version")
                .first()
            )
            next_version = (last_version.version if last_version else 0) + 1

            base_name, ext = os.path.splitext(file_name)
            versioned_filename = f"{base_name}_v{next_version}{ext}"

            archived_path = os.path.join(archive_dir, versioned_filename)
            deleted_path = os.path.join(deleted_dir, versioned_filename)

            if os.path.exists(file_path):
                shutil.move(file_path, archived_path)   # Move file to archive
                shutil.copy2(archived_path, deleted_path)  # Copy also into delete

            # 🔹 Update DB
            file_obj.version = next_version
            file_obj.file_name = versioned_filename
            file_obj.save()

            models.DeletedFileRecord.objects.create(
                data_file=file_obj).save()


            serializer = serializers.DataFileSerializer(file_obj)
            return Response(
                {
                    "success": 1,
                    "message": "File archived and copied to deleted folder with new version",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except DataFile.DoesNotExist:
            return Response(
                {"success": 0, "message": "File not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"success": 0, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class UploadStatusView(APIView):
    def get(self, request, upload_id):
        progress = UPLOAD_PROGRESS.get(upload_id, 0)
        print(progress,'.....................................................................................')
        return Response({"upload_id": upload_id, "progress": progress})


def sanitize_data(data):
    if isinstance(data, dict):
        return {k: sanitize_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data(i) for i in data]
    elif isinstance(data, float):
        return data if isfinite(data) else 0  # or None
    else:
        return data

class DataFileLatestView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, object_id):
        try:
            mode = request.query_params.get("mode", "view")  # default = view
            print("\n[DEBUG] Incoming GET request ----------------------")
            print("[DEBUG] object_id:", object_id)
            print("[DEBUG] mode:", mode)
            if mode == 'download':
                latest_file = (
                models.DataFile.objects.filter(id=object_id, version=0)
                .first()
                )
            else:
                latest_file = (
                    models.DataFile.objects.filter(data_object_id=object_id, version=0)
                    .first()
                )

            print("[DEBUG] latest_file object:", latest_file)

            if not latest_file:
                print("[DEBUG] No DataFile found for object_id:", object_id)
                return Response(
                    {"success": 0, "message": "No file found", "data": {}},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Serialize the latest file
            serializer = serializers.DataFileSerializer(latest_file)
            safe_data = sanitize_data(serializer.data)
            print("[DEBUG] Serialized DataFile:", safe_data)

            # Extract file path
            file_name = str(latest_file.file_name)
            print("[DEBUG] file_name:", file_name)

            object_name = file_name.split(".")[0]
            print("[DEBUG] object_name (from file_name):", object_name)

            base_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower())
            print("[DEBUG] base_dir:", base_dir)

            file_path = os.path.normpath(os.path.join(base_dir, file_name))
            print("[DEBUG] file_path:", file_path)

            # ----------------
            # Mode 1: VIEW
            # ----------------
            if mode == "view":
                print("[DEBUG] Entered VIEW mode")
                data = {}
                if file_path.endswith(".xlsx") and os.path.exists(file_path):
                    print("[DEBUG] File exists, opening with pandas:", file_path)
                    with pd.ExcelFile(file_path) as xls:
                        print("[DEBUG] Sheets found:", xls.sheet_names)
                        for sheet in xls.sheet_names:
                            df = xls.parse(sheet)
                            fields = df.columns.tolist()
                            top_rows = df.head(30).to_dict(orient="records")
                            safe_rows = sanitize_data(top_rows)
                            data[sheet] = {
                                "fields": fields,
                                "rows": safe_rows
                            }
                            print(f"[DEBUG] Processed sheet: {sheet}, rows: {len(df)}")

                else:
                    print("[DEBUG] File missing or not .xlsx:", file_path)

                return Response(
                    {
                        "success": 1,
                        "message": "Latest file data retrieved successfully",
                        "data": {
                            "file_info": safe_data,
                            "excel_preview": data
                        },
                    },
                    status=status.HTTP_200_OK,
                )

            # ----------------
            # Mode 2: DOWNLOAD
            # ----------------
            elif mode == "download":
                print("[DEBUG] Entered DOWNLOAD mode")
                print("[DEBUG] Checking file path:", file_path)

                if not os.path.exists(file_path):
                    print("[DEBUG] File not found at path:", file_path)
                    return Response(
                        {"success": 0, "message": "File not found", "data": {}},
                        status=status.HTTP_404_NOT_FOUND,
                    )

                print("[DEBUG] File exists, preparing Excel for download...")
                xls = pd.ExcelFile(file_path)
                print("[DEBUG] Sheets found:", xls.sheet_names)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    for sheet in xls.sheet_names:
                        df = xls.parse(sheet)
                        df.to_excel(writer, sheet_name=sheet, index=False)
                        print(f"[DEBUG] Copied sheet: {sheet}, rows: {len(df)}")

                    # Add mapping tab
                    print("[DEBUG] Fetching mapping specs for object_name:", object_name)
                    objectNameObj = models.DataObject.objects.filter(objectName=object_name).first()
                    print("[DEBUG] DataObject fetched:", objectNameObj)

                    if objectNameObj:
                        objectNameId = objectNameObj.id
                        spec_data = models.Specs.objects.filter(objectName=objectNameId).values()
                        spec_df = pd.DataFrame(spec_data)
                        print("[DEBUG] Spec rows:", len(spec_df))

                        spec_df = spec_df.drop(columns=["id", "position"], errors="ignore")
                        spec_df = spec_df.rename(columns={"objectName_id": "objectName"})
                        spec_df["objectName"] = object_name
                        spec_df.to_excel(writer, sheet_name="mapping", index=False)
                        print("[DEBUG] Added mapping tab")

                output.seek(0)
                print("[DEBUG] Returning Excel file as response")
                response = HttpResponse(
                    output,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="{file_name}"'
                return response

            else:
                print("[DEBUG] Invalid mode:", mode)
                return Response(
                    {"success": 0, "message": "Invalid mode", "data": {}},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:
            print("[ERROR] Exception in get():", str(e))
            import traceback; traceback.print_exc()
            return Response(
                {"success": 0, "message": f"Error: {str(e)}", "data": {}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class LatestValidatedFilesView(APIView):
    ''' Fetch latest validated (validation=1) file per DataObject '''
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    from django.db.models import Exists, OuterRef, Max

    def get(self, request):
        context = {
            "success": 0,
            "message": "Something went wrong.",
            "data": [],
        }
        try:
            # Step 1: Subquery to check if the data_object has any previous validation
            validated_subquery = models.DataFile.objects.filter(
                data_object=OuterRef('data_object'),
                validation=1
            )

            # Step 2: Fetch version=0 files with previous validation
            version0_files = models.DataFile.objects.filter(
                version=0
            ).annotate(
                has_previous_validation=Exists(validated_subquery)
            ).filter(has_previous_validation=True)

            # Step 3: Prepare a dict to track latest file per data_object
            latest_files_dict = {}

            # Add version=0 files first
            for file in version0_files:
                latest_files_dict[file.data_object] = file

            # Step 4: For data_objects without version=0 validated file, fetch latest validated file
            data_object_ids_with_v0 = version0_files.values_list('data_object', flat=True)
            fallback_files = (
                models.DataFile.objects
                .filter(validation=1)
                .exclude(data_object__in=data_object_ids_with_v0)
                .order_by('data_object', '-validated_at')
            )

            # Add fallback latest file per data_object
            for file in fallback_files:
                if file.data_object not in latest_files_dict:
                    latest_files_dict[file.data_object] = file

            # Step 5: Final list of files (unique per data_object)
            combined_files = list(latest_files_dict.values())

            serializer = serializers.DataFileSerializer(combined_files, many=True)

            if not serializer.data:
                context["message"] = "No validated files found."
                return Response(context, status=status.HTTP_404_NOT_FOUND)

            context["success"] = 1
            context["message"] = "Latest validated files fetched successfully."
            context["data"] = serializer.data

            return Response(context, status=status.HTTP_200_OK)

        except Exception as e:
            context["message"] = str(e)
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# Global in-memory store for progress objects
PROGRESS_STORE = {}

# //////1:52
def create_progress(task_name: str, data_object):
    task_id = str(uuid.uuid4())
    progress_obj = {
        "id": task_id,
        "task_name": task_name,
        "progress": 0,
        "message": "Started",
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    PROGRESS_STORE[task_id] = progress_obj

    # Create DB record
    models.ValidationProgress.objects.create(
        data_object=data_object,
        task_id=task_id,
        progress=0,
        status="pending"
    )

    print(f"[PROGRESS] Created tracker: {task_id} for task: {task_name}")
    return task_id, progress_obj

# 2:19
# def update_progress(task_id: str, progress: int, message: str = ""):
#     # Update memory
#     if task_id in PROGRESS_STORE:
#         PROGRESS_STORE[task_id]["progress"] = progress
#         if message:
#             PROGRESS_STORE[task_id]["message"] = message
#         PROGRESS_STORE[task_id]["updated_at"] = datetime.now().isoformat()

#     # Update DB
#     try:
#         vp = models.ValidationProgress.objects.get(task_id=task_id)
#         vp.progress = progress
#         if progress >= 100:
#             vp.status = "completed"
#         elif progress == 0 and "Error" in message:
#             vp.status = "failed"
#         else:
#             vp.status = "running"
#         vp.save(update_fields=["progress", "status", "updated_at"])
#     except models.ValidationProgress.DoesNotExist:
#         print(f"[DB WARNING] No ValidationProgress found for task {task_id}")

#     print(f"[PROGRESS] {task_id} -> {progress}% | {message}")

# new2
# def update_progress(task_id: str, progress: int, message: str = "", success: int | None = None):
#     # Update memory
#     if task_id in PROGRESS_STORE:
#         PROGRESS_STORE[task_id]["progress"] = progress
#         if message:
#             PROGRESS_STORE[task_id]["message"] = message
#         PROGRESS_STORE[task_id]["updated_at"] = datetime.now().isoformat()
#         # ✅ maintain success in memory
#         if success is not None:
#             PROGRESS_STORE[task_id]["success"] = success
#         elif progress == 0 and "Error" in message:
#             PROGRESS_STORE[task_id]["success"] = 0
#         elif progress >= 100:
#             PROGRESS_STORE[task_id]["success"] = 1
#         else:
#             PROGRESS_STORE[task_id].setdefault("success", 1)

#     # Update DB
#     try:
#         vp = models.ValidationProgress.objects.get(task_id=task_id)
#         vp.progress = progress
#         if success is not None:
#             vp.success = bool(success)
#         elif progress >= 100:
#             vp.status = "completed"
#             vp.success = True
#         elif progress == 0 and "Error" in message:
#             vp.status = "failed"
#             vp.success = False
#         else:
#             vp.status = "running"
#             if vp.success is None:
#                 vp.success = True
#         vp.save(update_fields=["progress", "status", "success", "updated_at"])
#     except models.ValidationProgress.DoesNotExist:
#         print(f"[DB WARNING] No ValidationProgress found for task {task_id}")

#     print(f"[PROGRESS] {task_id} -> {progress}% | {message}")

# new3
def update_progress(task_id: str, progress: int, message: str = "", success: int = None):
    # Update memory
    if task_id in PROGRESS_STORE:
        PROGRESS_STORE[task_id]["progress"] = progress
        if message:
            PROGRESS_STORE[task_id]["message"] = message
        PROGRESS_STORE[task_id]["updated_at"] = datetime.now().isoformat()
        if success is not None:
            PROGRESS_STORE[task_id]["success"] = success
        else:
            if progress >= 100:
                PROGRESS_STORE[task_id]["success"] = 1
            elif "Error" in message or "Missing" in message or "not found" in message:
                PROGRESS_STORE[task_id]["success"] = 0
            else:
                PROGRESS_STORE[task_id].setdefault("success", 1)

    # Update DB
    try:
        vp = models.ValidationProgress.objects.get(task_id=task_id)
        vp.progress = progress

        # Handle success/failure explicitly
        if progress >= 100 and (success is None or success == 1):
            vp.status = "completed"
            vp.success = True
        elif success == 0 or "Error" in message or "Missing" in message or "not found" in message:
            vp.status = "failed"
            vp.success = False
        else:
            vp.status = "running"
            if vp.success is None:
                vp.success = True

        vp.save(update_fields=["progress", "status", "success", "updated_at"])
    except models.ValidationProgress.DoesNotExist:
        print(f"[DB WARNING] No ValidationProgress found for task {task_id}")

    print(f"[PROGRESS] {task_id} -> {progress}% | {message} | success={PROGRESS_STORE.get(task_id, {}).get('success')}")

# new1
# def update_progress(task_id: str, progress: int, message: str = ""):
#     # Update memory
#     if task_id in PROGRESS_STORE:
#         PROGRESS_STORE[task_id]["progress"] = progress
#         if message:
#             PROGRESS_STORE[task_id]["message"] = message
#         PROGRESS_STORE[task_id]["updated_at"] = datetime.now().isoformat()
#         # ✅ maintain success in memory
#         if progress == 0 and "Error" in message:
#             PROGRESS_STORE[task_id]["success"] = 0
#         elif progress >= 100:
#             PROGRESS_STORE[task_id]["success"] = 1
#         else:
#             PROGRESS_STORE[task_id].setdefault("success", 1)

#     # Update DB
#     try:
#         vp = models.ValidationProgress.objects.get(task_id=task_id)
#         vp.progress = progress
#         if progress >= 100:
#             vp.status = "completed"
#             vp.success = True
#         elif progress == 0 and "Error" in message:
#             vp.status = "failed"
#             vp.success = False
#         else:
#             vp.status = "running"
#             # keep current success if already False
#             if vp.success is None:
#                 vp.success = True
#         vp.save(update_fields=["progress", "status", "success", "updated_at"])
#     except models.ValidationProgress.DoesNotExist:
#         print(f"[DB WARNING] No ValidationProgress found for task {task_id}")

#     print(f"[PROGRESS] {task_id} -> {progress}% | {message}")

# OLD
# def get_progress(task_id: str):
#     return PROGRESS_STORE.get(task_id, {"progress": 0, "message": ""})
def get_progress(task_id):
    """
    Fetch the latest progress for a given task_id.
    Always returns 'success' key (default=1).
    """
    progress = PROGRESS_STORE.get(task_id, {})
    return {
        "task_id": task_id,
        "progress": progress.get("progress", 0),
        "message": progress.get("message", "No progress yet"),
        "success": progress.get("success", 1)  # ✅ default to 1
    }

# OLD
# def run_validation_in_background(task_id, data_object_id, request_data):
#     """All heavy validation logic moved here."""
#     try:
#         update_progress(task_id, 5, "Fetching DataObject")
#         print(f"[Thread] Fetching DataObject {data_object_id}")
#         data_object = models.DataObject.objects.filter(id=data_object_id).first()
#         if not data_object:
#             update_progress(task_id, 6, "DataObject not found")
#             print("[Thread ERROR] DataObject not found")
#             return

#         # Check own file
#         update_progress(task_id, 6, "Checking main data file")
#         print("[Thread] Checking main data file")
#         own_file_exists = models.DataFile.objects.filter(data_object=data_object, version=0).exists()
#         if not own_file_exists:
#             update_progress(task_id, 7, f"Data file for '{data_object.objectName}' not found")
#             print(f"[Thread ERROR] Data file for '{data_object.objectName}' not found")
#             return

#         # Check dependencies
#         update_progress(task_id, 7, "Checking dependencies")
#         print("[Thread] Checking dependencies")
#         dependencies = data_object.dependencies or []
#         rules_applied_qs = models.RuleApplied.objects.filter(spec__objectName=data_object.id)
#         target_objects_dependencies = []
#         for rule in rules_applied_qs:
#             targets = get_target_specs(rule.rule_applied_data)
#             target_objects_dependencies.extend(targets)
#         dependencies = list(set(dependencies + target_objects_dependencies))
#         update_progress(task_id, 8, "verifying depenedency files..")

#         missing_files = []
#         for dep_name in dependencies:
#             dep_object = models.DataObject.objects.filter(objectName=dep_name).first()
#             if not dep_object or not models.DataFile.objects.filter(data_object=dep_object, version=0).exists():
#                 missing_files.append(dep_name)
#                 print(f"[Thread] Missing dependency: {dep_name}")
#         if missing_files:
#             update_progress(task_id, 9, f"Missing dependencies: {', '.join(missing_files)}")
#             print(f"[Thread ERROR] Missing dependencies: {', '.join(missing_files)}")
#             return

#         update_progress(task_id, 9, "Dependencies validated")
#         print("[Thread] Dependencies validated")

#         # Start default validations
#         update_progress(task_id, 10, "Running default validations")
#         paths = create_and_get_working_file_path(request_data.get("dataObjectId"))
#         print(f"[Thread] Working paths: {paths}")
#         resultLog1 = run_default_validators(
#             file_path=paths.get('working_file_path'),
#             log_file_path=paths.get('working_log_file_path'),
#             primary_field=request_data.get("fieldId"),
#             task_id=task_id,
#             update_progress_fun=update_progress
#         )
#         update_progress(task_id, 65, "default Validation completed successfully")
#         # delete_working_directory(paths.get('working_file_path'))
#         print("[Thread] Deleted working directory")

#         # Prepare logs
#         update_progress(task_id, 68, "Processing logs")
#         source_file = get_file_path_with_object_name(data_object.objectName)
#         working_log_file_path = paths.get('working_log_file_path')
#         try:
#             with open(working_log_file_path, "rb") as f:
#                 existing_log = pd.read_excel(f)
#             print(f"[Thread] Existing log loaded. Rows: {len(existing_log)}")
#         except FileNotFoundError:
#             existing_log = pd.DataFrame(columns=["primary_field", "rule_data", "time"])
#             print("[Thread] No existing log found, created empty log")

#         # Run custom rule validations
#         ###############################################
#         update_progress(task_id, 70, "Running custom rule validations")
#         new_logs_list = []
#         rules_applied_qs = models.RuleApplied.objects.filter(spec__objectName=data_object.id)
#         targets_obj = {}
#         total_rules = rules_applied_qs.count()

#         for i, rule in enumerate(rules_applied_qs):
#             targets = get_target_specs(json_spec=rule.rule_applied_data)
#             for obj in targets:
#                 targets_obj[obj] = get_file_path_with_object_name(obj)
#             print(f"[Thread] Running custom rule {rule.id} on targets: {targets_obj}")

#             resultLog2 = run_custom_rule_validation(
#                 rule_name=rule.rule_applied,
#                 json_spec=rule.rule_applied_data,
#                 source_file=source_file,
#                 target_files=targets_obj,
#                 rule_description=rule.description or ""
#             )
#             print(f"[Thread] Custom rule {rule.id} produced {len(resultLog2)} logs")
#             new_logs_list.append(pd.DataFrame(resultLog2, columns=["primary_field", "rule_data", "time"]))

#             # Update progress for this rule
#             if task_id and total_rules > 0:
#                 incremental_progress = int((i + 1) / total_rules * 25)  # 25% for this loop
#                 update_progress(task_id, 70 + incremental_progress, f"Running custom rule {i + 1}/{total_rules}")

#         ############
#         if new_logs_list:
#             all_new_logs = pd.concat(new_logs_list, ignore_index=True)
#             final_log = pd.concat([existing_log, all_new_logs], ignore_index=True)
#         else:
#             final_log = existing_log

#         with pd.ExcelWriter(working_log_file_path, engine="openpyxl", mode="w") as writer:
#             final_log.to_excel(writer, index=False)

#         actual_log_file_path = os.path.join(paths.get('logging_dir'), paths.get("log_file_name"))

#         # Copy working log file to actual log file
#         with pd.ExcelFile(working_log_file_path) as xls:
#             with pd.ExcelWriter(actual_log_file_path, engine="openpyxl") as writer:
#                 # Copy existing sheets
#                 for sheet in xls.sheet_names:
#                     df = xls.parse(sheet)
#                     df.to_excel(writer, sheet_name=sheet, index=False)

#         delete_working_directory(paths.get('working_file_path'))
#         update_progress(task_id, 98, "written in log file successfully !")
#         print("[Thread] Final log written to Excel")

#         # Update DataFile validation status
#         data_file = models.DataFile.objects.filter(data_object=data_object, version=0).first()
#         if data_file:
#             if not final_log.empty:
#                 # data_file.validation = models.DataFile.ValidationStatus.FAILED
#                 data_file.status = "VALIDATION COMPLETED WITH ERRORS"
#             else:
#                 # data_file.validation = models.DataFile.ValidationStatus.VALIDATED
#                 data_file.status = "VALIDATION COMPLETED WITH NO ERRORS"

#             data_file.validated_at = timezone.now()
#             data_file.validation = 1
#             data_file.save()
#             print("[Thread] DataFile validation updated")

#         update_progress(task_id, 100, "Validation completed successfully")
#         print("[Thread] Validation completed successfully")

#     except Exception as e:
#         update_progress(task_id, 0, f"Error: {str(e)}")
#         print(f"[Thread ERROR] {str(e)}")

def run_validation_in_background(task_id, data_object_id, request_data):
    """All heavy validation logic moved here."""
    try:
        update_progress(task_id, 5, "Fetching DataObject", success=1)
        print(f"[Thread] Fetching DataObject {data_object_id}")
        data_object = models.DataObject.objects.filter(id=data_object_id).first()
        if not data_object:
            update_progress(task_id, 6, "DataObject not found", success=0)
            print("[Thread ERROR] DataObject not found")
            return

        # Check own file
        update_progress(task_id, 6, "Checking main data file", success=1)
        print("[Thread] Checking main data file")
        own_file_exists = models.DataFile.objects.filter(data_object=data_object, version=0).exists()
        if not own_file_exists:
            update_progress(task_id, 7, f"Data file for '{data_object.objectName}' not found", success=0)
            print(f"[Thread ERROR] Data file for '{data_object.objectName}' not found")
            return

        # Check dependencies
        update_progress(task_id, 7, "Checking dependencies", success=1)
        print("[Thread] Checking dependencies")
        dependencies = data_object.dependencies or []
        rules_applied_qs = models.RuleApplied.objects.filter(spec__objectName=data_object.id)
        target_objects_dependencies = []
        for rule in rules_applied_qs:
            targets = get_target_specs(rule.rule_applied_data)
            target_objects_dependencies.extend(targets)
        dependencies = list(set(dependencies + target_objects_dependencies))
        update_progress(task_id, 8, "Verifying dependency files...", success=1)

        missing_files = []
        for dep_name in dependencies:
            dep_object = models.DataObject.objects.filter(objectName=dep_name).first()
            if not dep_object or not models.DataFile.objects.filter(data_object=dep_object, version=0).exists():
                missing_files.append(dep_name)
                print(f"[Thread] Missing dependency: {dep_name}")
        if missing_files:
            update_progress(task_id, 9, f"Missing dependencies: {', '.join(missing_files)}", success=0)
            print(f"[Thread ERROR] Missing dependencies: {', '.join(missing_files)}")
            return

        update_progress(task_id, 9, "Dependencies validated", success=1)
        print("[Thread] Dependencies validated")

        # Start default validations
        update_progress(task_id, 10, "Running default validations", success=1)
        paths = create_and_get_working_file_path(request_data.get("dataObjectId"))
        print(f"[Thread] Working paths: {paths}")
        resultLog1 = run_default_validators(
            file_path=paths.get('working_file_path'),
            log_file_path=paths.get('working_log_file_path'),
            primary_field=request_data.get("fieldId"),
            task_id=task_id,
            update_progress_fun=update_progress
        )
        update_progress(task_id, 65, "Default validation completed successfully", success=1)
        print("[Thread] Deleted working directory")

        # Prepare logs
        update_progress(task_id, 68, "Processing logs", success=1)
        source_file = get_file_path_with_object_name(data_object.objectName)
        working_log_file_path = paths.get('working_log_file_path')
        try:
            with open(working_log_file_path, "rb") as f:
                existing_log = pd.read_excel(f)
            print(f"[Thread] Existing log loaded. Rows: {len(existing_log)}")
        except FileNotFoundError:
            existing_log = pd.DataFrame(columns=["primary_field", "rule_data", "time"])
            print("[Thread] No existing log found, created empty log")

        # Run custom rule validations
        update_progress(task_id, 70, "Running custom rule validations", success=1)
        new_logs_list = []
        rules_applied_qs = models.RuleApplied.objects.filter(spec__objectName=data_object.id)
        targets_obj = {}
        total_rules = rules_applied_qs.count()

        for i, rule in enumerate(rules_applied_qs):
            targets = get_target_specs(json_spec=rule.rule_applied_data)
            for obj in targets:
                targets_obj[obj] = get_file_path_with_object_name(obj)
            print(f"[Thread] Running custom rule {rule.id} on targets: {targets_obj}")

            resultLog2 = run_custom_rule_validation(
                rule_name=rule.rule_applied,
                json_spec=rule.rule_applied_data,
                source_file=source_file,
                target_files=targets_obj,
                rule_description=rule.description or ""
            )
            print(f"[Thread] Custom rule {rule.id} produced {len(resultLog2)} logs")
            new_logs_list.append(pd.DataFrame(resultLog2, columns=["primary_field", "rule_data", "time"]))

            # Update progress for this rule
            if task_id and total_rules > 0:
                incremental_progress = int((i + 1) / total_rules * 25)  # 25% for this loop
                update_progress(task_id, 70 + incremental_progress, f"Running custom rule {i + 1}/{total_rules}", success=1)

        # Merge logs
        if new_logs_list:
            all_new_logs = pd.concat(new_logs_list, ignore_index=True)
            final_log = pd.concat([existing_log, all_new_logs], ignore_index=True)
        else:
            final_log = existing_log

        with pd.ExcelWriter(working_log_file_path, engine="openpyxl", mode="w") as writer:
            final_log.to_excel(writer, index=False)

        actual_log_file_path = os.path.join(paths.get('logging_dir'), paths.get("log_file_name"))

        # Copy working log file to actual log file
        with pd.ExcelFile(working_log_file_path) as xls:
            with pd.ExcelWriter(actual_log_file_path, engine="openpyxl") as writer:
                for sheet in xls.sheet_names:
                    df = xls.parse(sheet)
                    df.to_excel(writer, sheet_name=sheet, index=False)

        delete_working_directory(paths.get('working_file_path'))
        update_progress(task_id, 98, "Written in log file successfully!", success=1)
        print("[Thread] Final log written to Excel")

        # Update DataFile validation status
        data_file = models.DataFile.objects.filter(data_object=data_object, version=0).first()
        if data_file:
            if not final_log.empty:
                data_file.status = "VALIDATION COMPLETED WITH ERRORS"
            else:
                data_file.status = "VALIDATION COMPLETED WITH NO ERRORS"

            data_file.validated_at = timezone.now()
            data_file.validation = 1
            data_file.save()
            print("[Thread] DataFile validation updated")

        update_progress(task_id, 100, "Validation completed successfully", success=1)
        print("[Thread] Validation completed successfully")

    except Exception as e:
        update_progress(task_id, 0, f"Error: {str(e)}", success=0)
        print(f"[Thread ERROR] {str(e)}")


# 8:34///
class PreValidationCheckAndValidationView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def post(self, request, data_object_id):
        print(f"\n[API] Validation API called with data_object_id: {data_object_id}")

        # 🔹 Fetch the object
        data_object = get_object_or_404(models.DataObject, id=data_object_id)

        try:
        # 🔹 Check object-level permissions
            self.check_object_permissions(request, data_object)
        except PermissionDenied as e:
            return Response(
                {
                    "success": 0,
                    "message": str(e) or "You do not have permission to perform this action.",
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # 🔹 Step 1: Check if this object already has a running validation
        existing_in_progress = models.ValidationProgress.objects.filter(
        data_object_id=data_object_id, 
        status__in=["pending", "running"]
        ).exists()

        if existing_in_progress:
            return Response({
                "success": 0,
                "message": "Validation is already in progress for this DataObject. Please wait until it finishes."
            }, status=status.HTTP_400_BAD_REQUEST)

        task_id = str(uuid.uuid4())  # or uuid if you prefer

        # 🔹 Step 2: Create new progress tracker in DB
        progress = models.ValidationProgress.objects.create(
                data_object_id=data_object_id,
                task_id=task_id,   # store it in DB
                progress=0,
                status="pending"
            )

        
        PROGRESS_STORE[task_id] = {
            "progress": 0,
            "status": "pending",
            "message": "Started",
            "updated_at": datetime.now().isoformat(),
            "source": "db"
        }

        # 🔹 Step 3: Start background thread
        thread = threading.Thread(
            target=run_validation_in_background,
            args=(task_id, data_object_id, request.data),
            daemon=True
        )
        thread.start()

        # 🔹 Step 4: Return task_id immediately
        return Response({
            "success": 1,
            "message": "Validation started",
            "data": {"task_id": task_id}
        }, status=status.HTTP_200_OK)

# 1:56 /////
# class ValidationProgressView(APIView):
#     authentication_classes = [JWTAuthentication]
#     permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

#     def get(self, request, task_id):
#         # First check DB
#         try:
#             vp = models.ValidationProgress.objects.get(task_id=task_id)
#             data = {
#                 "task_id": vp.task_id,
#                 "progress": vp.progress,
#                 "status": vp.status,
#                 "updated_at": vp.updated_at.isoformat(),
#                 "source": "db"
#             }
#         except models.ValidationProgress.DoesNotExist:
#             # Fall back to memory store
#             data = PROGRESS_STORE.get(task_id, {
#                 "task_id": task_id,
#                 "progress": 0,
#                 "message": "Task not found",
#                 "source": "memory"
#             })

#         return Response({"success": 1, "data": data})

# new
class ValidationProgressView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated, DataObjectWriteLockPermission]

    def get(self, request, task_id):
        try:
            vp = models.ValidationProgress.objects.get(task_id=task_id)
            data = {
                "task_id": vp.task_id,
                "progress": vp.progress,
                "status": vp.status,
                "updated_at": vp.updated_at.isoformat(),
                "success": 1 if vp.success else 0,   # ✅ include DB success
                "source": "db",
            }
        except models.ValidationProgress.DoesNotExist:
            memory_data = PROGRESS_STORE.get(task_id, {
                "task_id": task_id,
                "progress": 0,
                "message": "Task not found",
                "success": 0,  # ✅ default fail in memory if not found
                "source": "memory",
            })
            data = memory_data

        return Response({"success": data.get("success", 0), "data": data})

class GetLatestLogDataView(APIView):
    """
    API to fetch top 50 rows of the latest log data for a given object_name.
    If `?download=1` is passed, returns the file for download.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, object_name):
        context = {
            "success": 1,
            "message": "Latest log data retrieved successfully",
            "data": []
        }
        try:
            # Log directory path
            log_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower(), "Log")
            if not os.path.exists(log_dir):
                context["success"] = 0
                context["message"] = f"No log directory found for {object_name}"
                return Response(context, status=status.HTTP_404_NOT_FOUND)

            # Get all Excel log files
            log_files = glob.glob(os.path.join(log_dir, "*.xlsx"))
            if not log_files:
                context["success"] = 0
                context["message"] = f"No log files found for {object_name}"
                return Response(context, status=status.HTTP_404_NOT_FOUND)

            # Pick the latest log file
            latest_file = max(log_files, key=os.path.getmtime)

            # ✅ If download flag is passed → return file
            if request.query_params.get("download") == "1":
                # Use streaming file response with proper file handling
                response = FileResponse(
                    open(latest_file, 'rb'),
                    as_attachment=True,
                    filename=os.path.basename(latest_file)
                )
                response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response['Content-Length'] = os.path.getsize(latest_file)
                return response

            # ✅ Otherwise → return JSON response (top 50 rows)
            df = pd.read_excel(latest_file)
            df = df.fillna("").head(50)
            context["data"] = df.to_dict(orient="records")

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(
            context,
            status=status.HTTP_200_OK if context["success"] else status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class DownloadLatestLogAPIView(APIView):
    """
    API to download the latest Excel log file for a given object_name.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, object_name):
        # Log directory path
        log_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower(), "Log")
        if not os.path.exists(log_dir):
            raise Http404(f"No log directory found for {object_name}")

        # Get all Excel log files
        log_files = glob.glob(os.path.join(log_dir, "*.xlsx"))
        if not log_files:
            raise Http404(f"No log files found for {object_name}")

        # Pick the latest log file
        latest_file = max(log_files, key=os.path.getmtime)
        file_name = os.path.basename(latest_file)

        # Create temporary copy
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_file_path = os.path.join(temp_dir, file_name)
            shutil.copy2(latest_file, temp_file_path)

            # Read temporary file as bytes
            with open(temp_file_path, "rb") as f:
                file_bytes = f.read()

            # Return as download
            response = HttpResponse(
                file_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{file_name}"'
            return response  # temporary file is deleted automatically

UPLOAD_PROGRESS = {}  # store progress in memory (use Redis/DB in production)


def process_file(upload_id, file):
    """
    Simulated file processing in background thread.
    Replace with your actual logic (save file, parse, validate, etc.).
    """
    for i in range(0, 101, 20):  # fake steps
        time.sleep(1)
        UPLOAD_PROGRESS[upload_id] = i


@csrf_exempt
def handle_file(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            return JsonResponse({"success": False, "message": "No file uploaded"})

        upload_id = str(uuid.uuid4())
        UPLOAD_PROGRESS[upload_id] = 0

        # run processing in background thread
        threading.Thread(target=process_file, args=(upload_id, file)).start()

        return JsonResponse({"success": True, "upload_id": upload_id})
    return JsonResponse({"success": False, "message": "Invalid method"})


def upload_status(request, upload_id):
    progress = UPLOAD_PROGRESS.get(upload_id, 0)
    return JsonResponse({"upload_id": upload_id, "progress": progress})

signer = TimestampSigner()

class RequestApprovalView(APIView):
    """
    API endpoint to handle approval requests:
    - approval → send approval email with attachments and dynamic link
    - cancel_approval → cancel existing approval (set status back to waiting)
    - cancel_progress → disable latest dynamic link (make it unusable)
    """

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def post(self, request, pk):
        try:
            print(f"➡️ Incoming approval request for file ID: {pk}")

            file_obj = get_object_or_404(models.DataFile, pk=pk)
            object_name = file_obj.data_object.objectName
            print(f"✅ File found: {file_obj.file_name}, Object: {object_name}")

            # Get flag from request
            flag = request.data.get("flag")
            print(f"📌 Flag received: {flag}")

            if not flag:
                print("❌ No flag provided in request.")
                return Response({"success": 0, "message": "Flag is required"}, status=status.HTTP_400_BAD_REQUEST)

            # ---------------------------------------------------
            # 1. REQUESTING_FOR_APPROVAL
            # ---------------------------------------------------
            if flag == "approval":
                print("➡️ Processing approval request...")

                if file_obj.validation == 0:
                    print("❌ File is not validated yet.")
                    return Response({"success": 0, "message": "File is not validated yet"})

                elif file_obj.approval_status == models.DataFile.ApprovalStatus.APPROVED or file_obj.release:
                    print("❌ File is already approved.")
                    return Response({"success": 0, "message": "File is already approved"})

                elif file_obj.request_progress == 1:
                    print("❌ File is already in Request Progress.")
                    return Response({"success": 0, "message": "File is already in Request Progress"})

                # Get file paths
                print("📂 Getting file paths...")
                data_file_path, log_file_path = get_file_paths(object_name, file_obj.file_name)

                # Generate token
                print("🔑 Generating approval token...")
                token = signer.sign(str(file_obj.id))
                approve_url = request.build_absolute_uri(reverse("approval-form", args=[token]))
                print(f"🔗 Approval URL: {approve_url}")

                # Send email
                print("📧 Sending approval email...")
                subject = f"Approval Request for Data Load {object_name} (Valid 24 hrs)"
                body = f"""
                    Hello Approver,

                    Please find attached the data load file for {object_name} for your review and approval.

                    Click the link below to approve or reject the data load:
                    {approve_url}

                    Note: This link will expire in 24 hours or once the action is taken.

                    Thank you,
                    Data Load Team
                    """

                email = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, ['sivakrishna.aar@gmail.com'])

                if os.path.exists(data_file_path):
                    print(f"📎 Attaching data file: {data_file_path}")
                    email.attach_file(data_file_path)
                if log_file_path and os.path.exists(log_file_path):
                    print(f"📎 Attaching log file: {log_file_path}")
                    email.attach_file(log_file_path)

                email.send()
                print("✅ Approval email sent successfully.")

                # Update DB
                file_obj.approval_token = token
                file_obj.request_progress = 1
                file_obj.approval_link_used = False
                file_obj.save()
                print("💾 File approval request state saved to DB.")

                return Response({"success": 1, "message": "Approval email sent successfully."}, status=status.HTTP_200_OK)

            # ---------------------------------------------------
            # 2. REQUEST_TO_CANCEL_APPROVAL
            # ---------------------------------------------------
            elif flag == "cancel_approval":
                print("➡️ Processing cancel approval request...")

                if file_obj.approval_status == models.DataFile.ApprovalStatus.APPROVED:
                    if file_obj.validation == 0:
                        print("❌ File is not validated yet.")
                        return Response({"success": 0, "message": "File is not validated yet"})

                    elif file_obj.request_progress == 1:
                        print("❌ File is already in Request Progress.")
                        return Response({"success": 0, "message": "File is already in Request Progress"})

                    elif not file_obj.release:
                        print("❌ File is not in Approved Status.")
                        return Response({"success": 0, "message": "File is not in Approved Status"})

                

                    # Get file paths
                    print("📂 Getting file paths...")
                    data_file_path, log_file_path = get_file_paths(object_name, file_obj.file_name)

                    # Generate release token
                    print("🔑 Generating release token...")
                    token = signer.sign(str(file_obj.id))
                    release_url = request.build_absolute_uri(reverse("approval-form", args=[token]))
                    print(f"🔗 Release URL: {release_url}")

                    # Send cancellation email
                    print("📧 Sending cancel approval email...")
                    subject = f"❌ Approval Cancelation Request for Data Load {object_name}"
                    body = f"""
                        Hello Approver,

                        The approval for data load file **{object_name}** has been requested to release or cancel the Approve.

                        File: {file_obj.file_name}
                        Cancelled At: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}

                        👉 Click the link below to mark this file as **READY TO UPLOAD**:
                        {release_url}

                        Note: This link will expire in 24 hours or once the action is taken.

                        Regards,  
                        Data Load Team
                    """

                    email = EmailMessage(subject, body, settings.DEFAULT_FROM_EMAIL, ['sivakrishna.aar@gmail.com'])

                    if os.path.exists(data_file_path):
                        print(f"📎 Attaching data file: {data_file_path}")
                        email.attach_file(data_file_path)
                    if log_file_path and os.path.exists(log_file_path):
                        print(f"📎 Attaching log file: {log_file_path}")
                        email.attach_file(log_file_path)

                    email.send()
                    print("✅ Cancel approval email sent successfully.")

                    # Save state
                    file_obj.approval_token = token
                    file_obj.request_progress = 1
                    file_obj.approval_link_used = False
                    file_obj.save()
                    print("💾 Cancel approval request state saved to DB.")

                    return Response({"success": 1, "message": "Approval cancel request sent with release link."}, status=status.HTTP_200_OK)

                else:
                    print("❌ File is not in approved state.")
                    return Response({"success": 0, "message": "File is not in approved state."}, status=status.HTTP_400_BAD_REQUEST)

            # ---------------------------------------------------
            # 3. TO_CANCEL_PROGRESS
            # ---------------------------------------------------
            elif flag == "cancel_progress":
                print("➡️ Cancelling approval progress...")
                
                file_obj.request_progress = 0
                file_obj.approval_link_used = True
                file_obj.approval_token = None
                file_obj.save()
                print("💾 Approval progress cancelled & link disabled.")

                return Response({"success": 1, "message": "Approval progress cancelled and link disabled."}, status=status.HTTP_200_OK)

            # ---------------------------------------------------
            # Invalid Flag
            # ---------------------------------------------------
            else:
                print("❌ Invalid flag provided.")
                return Response({"success": 0, "message": "Invalid flag provided."}, status=status.HTTP_400_BAD_REQUEST)

        except models.DataFile.DoesNotExist:
            print("❌ File not found in DB.")
            return Response({"success": 0, "message": "File not found."}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"💥 Exception occurred: {str(e)}")
            return Response({"success": 0, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def approval_form(request, token):
    try:
        file_id = signer.unsign(token, max_age=86400)  # 24hr expiry
        file_obj = DataFile.objects.get(pk=file_id)
    except SignatureExpired:
        return HttpResponseForbidden("⏳ This approval link has expired (time-based).")
    except BadSignature:
        return HttpResponseForbidden("❌ Invalid approval link.")

    if file_obj.approval_token != token:
        return HttpResponseForbidden("⚠️ This approval link is no longer valid (a newer link was issued).")

    if file_obj.approval_link_used:
        return HttpResponseForbidden("⚠️ This approval link has already been used.")

    if request.method == "POST":
        action = request.POST.get("action")  # approve / reject
        comment_text = request.POST.get("comment", "").strip()

        if not comment_text:
            return HttpResponseForbidden("⚠️ Comment is required.")

        # Track values for ApprovalComment
        action_type = None
        action_status = None

        # --- CASE 1: Initial approval (status = 0)
        if file_obj.approval_status == 0:
            if action == "approve":
                file_obj.approval_status = 1
                file_obj.release = True
                action_type = ApprovalComment.Action.REQUEST
                action_status = ApprovalComment.ActionStatus.APPROVED
            elif action == "reject":
                file_obj.approval_status = 2
                action_type = ApprovalComment.Action.REQUEST
                action_status = ApprovalComment.ActionStatus.REJECTED

        # --- CASE 2: Cancel approval (status = 1)
        elif file_obj.approval_status == 1:
            if action == "approve":
                file_obj.approval_status = 0
                file_obj.release = False
                action_type = ApprovalComment.Action.CANCEL
                action_status = ApprovalComment.ActionStatus.APPROVED
            elif action == "reject":
                file_obj.approval_status = 1
                action_type = ApprovalComment.Action.CANCEL
                action_status = ApprovalComment.ActionStatus.REJECTED

        # --- CASE 3: Rejected file (status = 2)
        elif file_obj.approval_status == 2:
            if action == "approve":
                file_obj.approval_status = 1
                file_obj.release = True
                action_type = ApprovalComment.Action.REQUEST
                action_status = ApprovalComment.ActionStatus.APPROVED
            elif action == "reject":
                file_obj.approval_status = 2
                action_type = ApprovalComment.Action.REQUEST
                action_status = ApprovalComment.ActionStatus.REJECTED

        # Save DataFile state
        file_obj.approval_link_used = True
        file_obj.request_progress = 0
        file_obj.approval_token = ''
        file_obj.approved_at = timezone.now()

        if comment_text:
            file_obj.approver_comment = comment_text
        file_obj.save()

        # Save comment with action + status
        ApprovalComment.objects.create(
            data_file=file_obj,
            comment=comment_text,
            action=action_type,
            action_status=action_status,
        )

        return redirect("approval-success")

    return render(request, "approval_form.html", {"file": file_obj})


def approval_success_view(request):
    return render(request, "approval_success.html")

class DataObjectCommentsView(APIView):
    """
    API to fetch all comments for a given data_object_id
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, data_object_id):
        try:
            files = models.DataFile.objects.filter(data_object_id=data_object_id)
            if not files.exists():
                return Response(
                    {"success": 0, "message": "No files found for this data object."},
                    status=status.HTTP_404_NOT_FOUND,
                )

            comments = (
                models.ApprovalComment.objects.filter(data_file__in=files)
                .select_related("data_file")
                .order_by("-created_at")
            )

            data = [
                {
                    "id": c.id,
                    "file_id": c.data_file.id,
                    # "file_name": c.data_file.file_name,
                    "action": getattr(c, "action", None),          # ✅ new
                    "action_status": getattr(c, "action_status", None),  # ✅ new
                    "comment": c.comment,
                    "created_at": c.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                }
                for c in comments
            ]

            return Response({"success": 1, "data": data}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"success": 0, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

class SpecsDownloadUploadView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated,DataObjectWriteLockPermission]

    def get(self, request, *args, **kwargs):
        object_name = request.GET.get("objectName")
        if not object_name:
            return Response({"success": 0, "message": "❌ objectName is required.", "data": {}}, status=400)

        # ✅ fetch DataObject
        data_object = get_object_or_404(DataObject, objectName=object_name)

        # ✅ fetch specs linked to that object
        specs_qs = Specs.objects.filter(objectName=data_object).order_by("tab", "position")

        fields = [
            "company",
            "objectName",
            "tab",
            "field_id",
            "mandatory",
            "allowed_values",
            "sap_table",
            "sap_field_id",
            "sap_description",
        ]

        data = []
        if specs_qs.exists():
            for spec in specs_qs:
                data.append({
                    "company": spec.company,
                    "objectName": spec.objectName.objectName,
                    "tab": spec.tab,
                    "field_id": spec.field_id,
                    # 🔹 if mandatory = "No", keep it blank in Excel
                    "mandatory": "" if str(spec.mandatory).strip().lower() in ["no", "n"] else "Yes",
                    "allowed_values": ",".join(spec.allowed_values) if spec.allowed_values else "",
                    "sap_table": spec.sap_table or "",
                    "sap_field_id": spec.sap_field_id or "",
                    "sap_description": spec.sap_description or "",
                })
        else:
            data.append({
                "company": data_object.company,
                "objectName": data_object.objectName,
                "tab": "",
                "field_id": "",
                "mandatory": "",
                "allowed_values": "",
                "sap_table": "",
                "sap_field_id": "",
                "sap_description": "",
            })

        df = pd.DataFrame(data, columns=fields)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{object_name}_specs.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        with pd.ExcelWriter(response, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Specs")

        return response

    def post(self, request, *args, **kwargs):
        print("📥 Entered POST method for specs upload")
        file = request.FILES.get("file")
        if not file:
            print("❌ No file uploaded")
            return Response(
                {"success": 0, "message": "No file uploaded.", "data": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            print("📑 Reading Excel file...")
            df = pd.read_excel(file)
            print(f"✅ Excel loaded with {len(df)} rows and {len(df.columns)} columns")

            required_columns = [
                "objectName", "tab", "field_id", "mandatory",
                "allowed_values", "sap_table", "sap_field_id", "sap_description"
            ]
            for col in required_columns:
                if col not in df.columns:
                    print(f"❌ Missing column: {col}")
                    return Response(
                        {"success": 0, "message": f"Missing column: {col}", "data": {}},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            print("🔍 Pre-validating mandatory fields...")
            mandatory_fields = ["objectName", "tab", "field_id"]

            for idx, row in df.iterrows():
                for field in mandatory_fields:
                    if pd.isna(row.get(field)) or str(row.get(field)).strip() == "":
                        print(f"❌ Row {idx+2}: Missing field {field}")
                        return Response(
                            {
                                "success": 0,
                                "message": f"Row {idx + 2}: Missing required field '{field}'",
                                "data": {},
                            },
                            status=status.HTTP_400_BAD_REQUEST,
                        )

            print("🧹 Cleaning up rows...")
            df["objectName"] = df["objectName"].astype(str).str.strip()
            df["tab"] = df["tab"].astype(str).str.strip().str.lower()
            df["field_id"] = df["field_id"].astype(str).str.strip().str.upper()

            df["mandatory"] = df["mandatory"].apply(
                lambda val: "Yes" if str(val).strip().lower() in ["yes", "y"] else "No"
            )

            object_name = df["objectName"].iloc[0]
            if not all(df["objectName"] == object_name):
                print("❌ Inconsistent objectName values detected")
                return Response(
                    {
                        "success": 0,
                        "message": "All rows must have the same objectName.",
                        "data": {},
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            print(f"🔎 Fetching DataObject for objectName={object_name}")
            data_object = get_object_or_404(models.DataObject, objectName__iexact=object_name)

            try:
                print("🔐 Checking object-level permissions...")
                self.check_object_permissions(request, data_object)
            except PermissionDenied as e:
                print(f"❌ Permission denied: {str(e)}")
                return Response(
                    {
                        "success": 0,
                        "message": str(e) or "You do not have permission to perform this action.",
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            print("📂 Fetching existing specs...")
            existing_specs = Specs.objects.filter(objectName=data_object)
            existing_map = {(s.tab.lower(), s.field_id.upper()): s for s in existing_specs}
            print(f"✅ Found {len(existing_specs)} existing specs")

            uploaded_combinations = set()
            inserted, updated, deleted = 0, 0, 0

            print("⚡ Starting transaction for upsert + delete...")
            with transaction.atomic():
                for _, row in df.iterrows():
                    combo = (row["tab"], row["field_id"])
                    uploaded_combinations.add(combo)

                    allowed_vals_cell = row.get("allowed_values")
                    allowed_values = (
                        [v.strip() for v in str(allowed_vals_cell).split(",") if v.strip()]
                        if pd.notna(allowed_vals_cell) and str(allowed_vals_cell).strip() != ""
                        else None
                    )

                    if combo in existing_map:
                        print(f"✏️ Updating spec: {combo}")
                        spec = existing_map[combo]
                        spec.mandatory = row["mandatory"]
                        spec.allowed_values = allowed_values
                        spec.sap_table = (
                            row["sap_table"]
                            if pd.notna(row.get("sap_table")) and str(row.get("sap_table")).strip()
                            else None
                        )
                        spec.sap_field_id = (
                            row["sap_field_id"]
                            if pd.notna(row.get("sap_field_id")) and str(row.get("sap_field_id")).strip()
                            else None
                        )
                        spec.sap_description = (
                            row["sap_description"]
                            if pd.notna(row.get("sap_description")) and str(row.get("sap_description")).strip()
                            else None
                        )
                        spec.save()
                        updated += 1
                    else:
                        print(f"➕ Inserting new spec: {combo}")
                        Specs.objects.create(
                            company=row.get("company", 'GenOne'),  # keep original company if column exists
                            objectName=data_object,
                            tab=row["tab"],
                            field_id=row["field_id"],
                            mandatory=row["mandatory"],
                            allowed_values=allowed_values,
                            sap_table=row["sap_table"]
                            if pd.notna(row.get("sap_table")) and str(row.get("sap_table")).strip()
                            else None,
                            sap_field_id=row["sap_field_id"]
                            if pd.notna(row.get("sap_field_id")) and str(row.get("sap_field_id")).strip()
                            else None,
                            sap_description=row["sap_description"]
                            if pd.notna(row.get("sap_description")) and str(row.get("sap_description")).strip()
                            else None,
                        )
                        inserted += 1

                print("🗑️ Checking for deleted specs...")
                for combo, spec in existing_map.items():
                    if combo not in uploaded_combinations:
                        print(f"🗑️ Deleting spec: {combo}")
                        spec.ruleapplied_set.all().delete()
                        spec.delete()
                        deleted += 1

            print(f"✅ Upload completed: inserted={inserted}, updated={updated}, deleted={deleted}")
            return Response(
                {
                    "success": 1,
                    "message": "Specs uploaded successfully.",
                    "data": {"inserted": inserted, "updated": updated, "deleted": deleted},
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            print(f"❌ Exception occurred: {str(e)}")
            return Response(
                {"success": 0, "message": str(e), "data": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )


import pandas as pd
import sqlalchemy
import os
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


class DataLoad(APIView):
    """
    API to load and transform Excel migration data for a given objectName,
    then insert into MySQL tables.
    """

    def get(self, request, *args, **kwargs):
        print("➡️ Entered DataLoad GET method")

        object_name_id = request.GET.get("objectName")
        if not object_name_id:
            return Response(
                {"success": 0, "message": "❌ objectName is required.", "data": {}},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # ----------------------------
            # STEP 0: Get Object Name
            # ----------------------------
            object_name_qry = models.DataObject.objects.filter(id=object_name_id).first()
            object_name = object_name_qry.objectName if object_name_qry else None
            print(f"🔍 Object Name Resolved: {object_name}")

            # ✅ Check release status before proceeding
            file_record = models.DataFile.objects.filter(data_object_id=object_name_id).last()
            if not file_record or not file_record.release:
                print(f"⛔ Data load blocked: Object {object_name} not released")
                return Response(
                    {
                        "success": 0,
                        "message": f"⛔ Data load blocked: {object_name} is not released yet",
                        "data": {},
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )

            working_info = create_and_get_working_file_path(object_name_id)
            if not working_info or not os.path.exists(working_info["working_file_path"]):
                print(f"❌ Working file not found for object: {object_name}")
                return Response(
                    {"success": 0, "message": f"❌ Working file not found for {object_name}", "data": {}},
                    status=status.HTTP_404_NOT_FOUND,
                )

            file_path = working_info["working_file_path"]
            print(f"📂 Using working file: {file_path}")

            # ----------------------------
            # STEP 1: Load Excel + Build Mapping
            # ----------------------------
            print("📑 Reading Excel...")
            with pd.ExcelFile(file_path) as xls:
                mapping_df = pd.read_excel(xls, sheet_name="mapping")
                print("✅ Mapping sheet loaded")
                print(mapping_df.head())

                tab_table_map = {}
                print("🔄 Building mapping dictionary...")
                for _, row in mapping_df.iterrows():
                    tab = row["tab"]
                    src_field = row["field_id"]
                    target_table = row["sap_table"]
                    target_field = row["sap_field_id"]

                    if pd.isna(target_table) or pd.isna(target_field):
                        print(f"⚠️ Skipping unmapped row: {row.to_dict()}")
                        continue

                    if tab not in tab_table_map:
                        tab_table_map[tab] = {}

                    if target_table not in tab_table_map[tab]:
                        tab_table_map[tab][target_table] = {}

                    tab_table_map[tab][target_table][src_field] = target_field

                print("✅ Mapping dictionary created")
                print(tab_table_map)

                # ----------------------------
                # STEP 2: Process Tabs
                # ----------------------------
                staging_data = {}
                print("🔄 Processing sheet tabs...")
                for tab_name in xls.sheet_names:
                    if tab_name == "mapping":
                        continue

                    print(f"📑 Processing Tab: {tab_name}")
                    df = pd.read_excel(xls, sheet_name=tab_name)
                    print(f"   ➡️ Loaded {len(df)} rows from {tab_name}")

                    if tab_name not in tab_table_map:
                        print(f"⚠️ No mapping found for tab {tab_name}, skipping...")
                        continue

                    for target_table, field_map in tab_table_map[tab_name].items():
                        mapped_df = pd.DataFrame()
                        print(f"   🔄 Mapping for Target Table: {target_table}")

                        for src_field, target_field in field_map.items():
                            if src_field in df.columns:
                                mapped_df[target_field] = df[src_field]
                                print(f"      ✅ Mapped {src_field} → {target_field}")
                            else:
                                pass
                                print(f"      ⚠️ Source field {src_field} not found in tab {tab_name}")

                        if target_table not in staging_data:
                            staging_data[target_table] = mapped_df
                        else:
                            staging_data[target_table] = pd.concat(
                                [staging_data[target_table], mapped_df],
                                ignore_index=True,
                            )

                print("✅ All tabs processed")
                print("📊 Staging Data Summary:")
                for k, v in staging_data.items():
                    pass
                    print(f"   - {k}: {len(v)} rows")

            # ----------------------------
            # STEP 3: Insert Into Database
            # ----------------------------
            print("🔌 Connecting to SAP HANA...")
            engine = sqlalchemy.create_engine(
                    f"mysql+pymysql://{SAP_DB_USER}:{SAP_DB_PASS}@{SAP_DB_HOST}:{SAP_DB_PORT}/{SAP_DB_NAME}"
                )
            print("✅ DB Connection established")

            result_summary = {}
            for table, df in staging_data.items():
                try:
                    if df.empty:
                        print(f"⚠️ Skipping {table} (0 rows)")
                        continue

                    print(f"📥 Inserting {len(df)} rows into {table}...")
                    df.to_sql(table, con=engine, if_exists="append", index=False)
                    print(f"✅ Inserted {len(df)} rows into {table}")
                    result_summary[table] = len(df)

                except Exception as e:
                    err_str = str(e)
                    print(f"❌ Error while inserting into {table}: {err_str}")

                    if "Duplicate entry" in err_str:
                        msg = f"❌ Duplicate entry error while inserting into {table} (Primary Key violation)"
                    elif "Unknown column" in err_str:
                        msg = f"❌ Column mismatch while inserting into {table} (check mapping vs DB schema)"
                    elif "doesn't exist" in err_str or "not found" in err_str.lower():
                        msg = f"❌ Target table {table} does not exist in database"
                    else:
                        msg = f"❌ Unexpected error while inserting into {table}: {err_str}"

                    return Response({"success": 0, "message": msg, "data": {}}, status=500)

            print("✅ Data Load Completed Successfully")
            print("📊 Final Insert Summary:", result_summary)

            # ----------------------------
            # STEP 4: Update data_load flag
            # ----------------------------
            try:
                file_record = models.DataFile.objects.filter(data_object_id=object_name_id).last()
                if file_record:
                    file_record.data_load = 1
                    file_record.save(update_fields=["data_load"])
                    print(f"✅ Updated data_load=1 for file: {file_record.file_name}")
                else:
                    print("⚠️ No DataFile found to update data_load")
            except Exception as e:
                print(f"⚠️ Failed to update data_load flag: {str(e)}")

            return Response(
                {
                    "success": 1,
                    "message": f"✅ Data load completed for {object_name}",
                    "data": {"staging_summary": result_summary},
                },
                status=200,
            )


        except Exception as e:
            print(f"💥 Critical Error: {str(e)}")
            return Response(
                {"success": 0, "message": f"❌ Error: {str(e)}", "data": {}},
                status=500,
            )



