
# views.py
import os
import io
import shutil
import pandas as pd
import django_filters
from collections import defaultdict

from django.conf import settings
from django.db import IntegrityError
from django.db.models import Q
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import get_object_or_404
from django.http import HttpResponse, JsonResponse

from rest_framework import status, generics, viewsets, filters
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action

from rest_framework_simplejwt.authentication import JWTAuthentication

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from . import models, serializers, messages, validators, constants
from .exceptions import SerializerError
from .models import DataObject, DataFile, Specs
from .serializers import FileUploadSerializer
from .pagination import StandardResultsSetPagination



class RuleAppliedFilter(django_filters.FilterSet):
    # create a custom alias for the long FK chain
    objectName = django_filters.CharFilter(
        field_name="spec__objectName__objectName", lookup_expr="exact"
    )

     # ✅ New filter for spec_id
    spec_id = django_filters.NumberFilter(
        field_name="spec__id", lookup_expr="exact"
    )

    created_after = django_filters.DateFilter(
        field_name="created_at", lookup_expr="gte"
    )
    created_before = django_filters.DateFilter(
        field_name="created_at", lookup_expr="lte"
    )

    class Meta:
        model = models.RuleApplied
        fields = ["objectName", "rule_applied", "created_after", "created_before"]

class RuleAppliedGetView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        queryset = models.RuleApplied.objects.select_related("spec").all()

        # Apply filters manually
        filterset = RuleAppliedFilter(request.GET, queryset=queryset)
        if not filterset.is_valid():
            return Response(filterset.errors, status=status.HTTP_400_BAD_REQUEST)
        queryset = filterset.qs


        # Apply ordering
        ordering = request.GET.get("ordering")
        if ordering:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by("-created_at")  # default ordering

        # ✅ Apply pagination
        paginator = StandardResultsSetPagination()
        paginated_qs = paginator.paginate_queryset(queryset, request, view=self)

        serializer = serializers.RuleAppliedTableSerializer(paginated_qs, many=True)
        return paginator.get_paginated_response(serializer.data)


class AllDataObjectView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        context = {
            "success": 1,
            "message": messages.DATA_FOUND,
            "data": {}
        }
        try:
            modules = models.DataObject.objects.all()
            module_ser = serializers.DataObjectSerializer(modules, many=True)
            context['data'] = module_ser.data
        except Exception as e:
            context['success'] = 0
            context['message'] = str(e)
        return Response(context)

class DataObjectView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    # pagination_class = CustomPagination  # Replace with your actual pagination class if different

    def get(self, request, id):
        context = {
            "success": 1,
            "message": messages.DATA_FOUND,
            "data": [],
        }
        try:
            
            record_id = id
            print('iddddddddddddd',id)
            # Single record fetch
            module = models.DataObject.objects.get(id=record_id)
            module_ser = serializers.DataObjectSerializer(module)
            context["data"] = module_ser.data
            context["count"] = 1

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def post(self, request, *args, **kwargs):
        context = {
            "success": 1,
            "message": messages.DATA_SAVED,
            "data": {},
        }
        try:
            validator = validators.ObjectDataValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            req_params = validator.validated_data
            module_instance = models.DataObject(**req_params)
            module_instance.clean()
            module_instance.save()

            context["data"] = serializers.DataObjectSerializer(module_instance).data

        except IntegrityError as e:
            context["success"] = 0
            if "Duplicate entry" in str(e):
                # Extract field causing error
                err_msg = str(e)
                duplicate_value = err_msg.split("'")[1]
                field_name = err_msg.split("for key")[1].split("'")[1]
                context["message"] = f"❌ '{duplicate_value}' already exists."
            else:
                context["message"] = "❌ Database Integrity Error."

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def put(self, request, id):
        context = {
            "success": 1,
            "message": messages.DATA_UPDATED,
            "data": {},
        }
        try:
            data = request.data
            record_id = id
            if not record_id:
                raise Exception("ID is required for update")
            
            validator = validators.ObjectDataValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)
            
            module_obj = models.DataObject.objects.get(id=record_id)
            for field in ["company", "dependencies", "module", "objectName"]:
                if field in data:
                    setattr(module_obj, field, data[field])
            
            # module_obj.full_clean()
            # print(request.data,'<<<<<<<<<--------')
            module_obj.save()

            context["data"] = serializers.DataObjectSerializer(module_obj).data

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def patch(self, request, *args, **kwargs):
        context = {
            "success": 1,
            "message": messages.DATA_UPDATED,
            "data": {},
        }
        try:
            data = request.data
            record_id = data.get("id")
            if not record_id:
                raise Exception("ID is required for partial update")

            module_obj = models.DataObject.objects.get(id=record_id)

            for field in ["company", "dependencies", "module", "objectName"]:
                if field in data:
                    setattr(module_obj, field, data[field])

            module_obj.full_clean()
            module_obj.save()

            context["data"] = serializers.DataObjectSerializer(module_obj).data

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

    def delete(self, request, id):
        context = {
            "success": 1,
            "message": messages.DATA_DELETED,
            "data": {},
        }
        try:
            if not id:
                raise Exception("ID is required for delete")

            module_obj = models.DataObject.objects.get(id=id)

            # 🔍 Check if this object is a dependency in other objects
            dependent_objects = models.DataObject.objects.filter(dependencies__contains=[module_obj.objectName])

            if dependent_objects.exists():
                raise Exception(
                    f"Cannot delete '{module_obj.objectName}' because it is a dependency of other objects."
                )

            module_obj.delete()

        except models.DataObject.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)



class AllDependenciesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        context = {
            "success": 1,
            "message": messages.DATA_FOUND,
            "data": []
        }
        try:
            # Collect all objectName values
            dependencies = list(models.DataObject.objects.values_list("objectName", flat=True))

            # Remove duplicates & sort
            unique_dependencies = sorted(set(dependencies))

            context["data"] = unique_dependencies

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)



class SpecsAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]
    
    def get(self, request, *args, **kwargs):
        context = {"success": 1, "message": messages.DATA_FOUND, "data": {}}
        try:
            object_id = kwargs.get("id")   # Get objectName_id from URL
            if not object_id:
                context["success"] = 0
                context["message"] = "objectName_id is required"
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

            # Filter only that objectName_id and order by tab + position
            specs = (
                models.Specs.objects
                .select_related("objectName")
                .filter(objectName_id=object_id)
                .order_by("tab", "position")
            )
            # print(models.DataObject.objects.get(id=object_id).objectName)

            if not specs.exists():
                context["success"] = 1
                context["message"] = "No specs data exists"
                context["data"] = {
                    'objectName':models.DataObject.objects.get(id=object_id).objectName
                }
                return Response(context, status=status.HTTP_200_OK)

            grouped_data = defaultdict(list)
            for spec in specs:
                grouped_data[spec.tab].append({
                    "id": spec.id,
                    "company": spec.company,
                    "field_id": (spec.field_id).upper(), # to upper case
                    "mandatory": spec.mandatory,
                    "allowed_values": spec.allowed_values,
                    "sap_table": spec.sap_table,
                    "sap_field_id": spec.sap_field_id,
                    "sap_description": spec.sap_description,
                    "position": spec.position,   # include position
                })

            tab_data = [
                {"tab": (tab_name).lower(), "fields": fields}
                for tab_name, fields in grouped_data.items()
            ]

            response_data = {
                "objectName": specs.first().objectName.objectName,
                "tabs": tab_data,
            }

            context["success"] = 1
            context["message"] = "Data fetched successfully"
            context["data"] = response_data
            return Response(context, status=status.HTTP_200_OK)

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            context["data"] = {}
            return Response(context, status=status.HTTP_400_BAD_REQUEST)
        
    def post(self, request, *args, **kwargs):
        context = {"success": 1, "message": "Data saved successfully", "data": {}}
        try:
            validator = validators.SpecsValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            req_params = validator.validated_data

            # Convert Yes/No to Boolean
            # req_params["mandatory"] = True if req_params["mandatory"] == "Yes" else False

            # Convert objectName ID → DataObject instance
            data_object = get_object_or_404(DataObject, pk=req_params["objectName"])
            req_params["objectName"] = data_object  

            # ✅ Auto-assign position based on tab
            existing_fields = models.Specs.objects.filter(
                objectName=data_object, tab=req_params["tab"]
            ).order_by("position")

            if existing_fields.exists():
                # Get last field position and increment
                last_position = existing_fields.last().position or 0
                req_params["position"] = last_position + 1
            else:
                # No fields in this tab → start at 0
                req_params["position"] = 0

            # Create new Specs record
            spec_instance = models.Specs.objects.create(**req_params)
            print(req_params,'--------------->>>>>>')
            context["data"] = serializers.SpecsSerializer(spec_instance).data
            

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
        return Response(context)

    def put(self, request, id=None, *args, **kwargs):
        context = {"success": 1, "message": "Data updated successfully", "data": {}}
        try:
            # ✅ validate incoming data"
            validator = validators.SpecsUpdateValidator(data=request.data)
            if not validator.is_valid():
                raise SerializerError(validator.errors)

            req_params = validator.validated_data

            # ✅ convert Yes/No to Boolean (if coming as text)"
            # if isinstance(req_params.get("mandatory"), str):
            #     req_params["mandatory"] = True if req_params["mandatory"] == "Yes" else False

            # ✅ fetch the Specs record at respective id--",id
            spec_instance = get_object_or_404(models.Specs, pk=id)

            # ✅ update objectName foreign key if passed"
            if "objectName" in req_params:
                data_object = get_object_or_404(DataObject, pk=req_params["objectName"])
                req_params["objectName"] = data_object

            # ✅ handle allowed_values (list or string)"
            if "allowed_values" in req_params:
                if isinstance(req_params["allowed_values"], str):
                    req_params["allowed_values"] = [
                        v.strip() for v in req_params["allowed_values"].split(",") if v.strip()
                    ]

            # ✅ update fields"
            for key, value in req_params.items():
                setattr(spec_instance, key, value)
            spec_instance.save()

            # ✅ return updated record"
            context["data"] = serializers.SpecsSerializer(spec_instance).data

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
        return Response(context)
    
    def delete(self, request,id=None):
        context = {
            "success": 1,
            "message": messages.DATA_DELETED,
            "data": {},
        }
        try:
            record_id = id
            if not record_id:
                raise Exception("ID is required for delete")

            module_obj = models.Specs.objects.get(id=record_id)
            module_obj.delete()

        except models.Specs.DoesNotExist:
            context["success"] = 0
            context["message"] = "Record not found"
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)


class ReorderSpecsView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def put(self, request, *args, **kwargs):
        try:
            object_id = request.data.get("objectName_id")
            tab = request.data.get("tab")
            fields = request.data.get("fields", [])

            if not object_id or not tab or not fields:
                return Response({"success": 0, "message": "Invalid data"}, status=status.HTTP_400_BAD_REQUEST)

            # update positions
            for field in fields:
                field_id = field.get("id")
                position = field.get("position")
                if field_id is not None and position is not None:
                    models.Specs.objects.filter(
                        id=field_id,
                        objectName_id=object_id,
                        tab=tab
                    ).update(position=position)

            return Response({"success": 1, "message": "Reordering saved successfully"}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"success": 0, "message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




class SpecsTemplateDownloadAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        context = {"success": 1, "message": "Template generated successfully", "data": {}}
        try:
            # ✅ Create workbook & sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Specs Template"

            # ✅ Define Specs model fields (exclude id/auto fields if required)
            fields = [
                "company",
                "objectName_id",
                "tab",
                "field_id",   
                "mandatory",        # Yes / No
                "allowed_values",   # Comma separated            
                "sap_table",
                "sap_field_id",
                "sap_description",
            ]

            # ✅ Write header row
            ws.append(fields)

            # ✅ Add dropdown validation for "mandatory" column
            # Column H = 8th column (mandatory field)
            dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
            ws.add_data_validation(dv)

            # Apply dropdown for a reasonable number of rows (say 1000)
            dv.add(f"E2:E1000")

            # (Optional) Example row for user guidance
            # ws.append([
            #     "f101", "38", "GenOne", "tab1",
            #     "t11", "s11", "desc...", "Yes",
            #     "11,22,33", "1",
            # ])

            # ✅ Prepare Excel response
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="SpecsTemplate.xlsx"'
            wb.save(response)

            return response

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            return JsonResponse(context, status=500)


class SpecsTemplateUploadAPIView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        context = {"success": 1, "message": "Data uploaded successfully", "data": []}
        try:
            file_obj = request.FILES.get("file")
            if not file_obj:
                raise Exception("No file uploaded")

            wb = openpyxl.load_workbook(file_obj)
            ws = wb.active

            # ✅ Expected header row
            expected_headers = [
                "field_id",
                "objectName_id",
                "company",
                "tab",
                "sap_table",
                "sap_field_id",
                "sap_description",
                "mandatory",
                "allowed_values",
                "position",
            ]

            headers = [cell.value for cell in ws[1]]
            if headers != expected_headers:
                raise Exception("Invalid template format. Please use the downloaded template.")

            # ✅ Iterate rows and save
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # skip empty rows
                    continue

                (
                    field_id,
                    objectName_id,
                    company,
                    tab,
                    sap_table,
                    sap_field_id,
                    sap_description,
                    mandatory,
                    allowed_values,
                    position,
                ) = row

                # ✅ handle mandatory Yes/No → Boolean
                mandatory_val = True if str(mandatory).lower() == "yes" else False

                # ✅ handle allowed_values → list
                allowed_list = (
                    [v.strip() for v in str(allowed_values).split(",") if v.strip()]
                    if allowed_values
                    else None
                )

                # ✅ ensure DataObject exists
                try:
                    data_object = models.DataObject.objects.get(objectName=objectName_id)
                except ObjectDoesNotExist:
                    data_object = models.DataObject.objects.create(objectName=objectName_id)

                # ✅ create or update Specs
                spec_obj, created = models.Specs.objects.update_or_create(
                    field_id=field_id,
                    defaults={
                        "objectName": data_object,
                        "company": company,
                        "tab": tab,
                        "sap_table": sap_table,
                        "sap_field_id": sap_field_id,
                        "sap_description": sap_description,
                        "mandatory": mandatory_val,
                        "allowed_values": allowed_list,
                        "position": position,
                    },
                )

                context["data"].append(serializers.SpecsSerializer(spec_obj).data)

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return JsonResponse(context)


class CustomRuleTemplateUIViewSet(viewsets.ModelViewSet):
    """
    Provides: list, retrieve (by pk), create, update, partial_update, destroy
    Extra action: retrieve_by_rule (GET /api/rules/schema/<rule_name>/)
    """

    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    queryset = models.CustomRuleTemplateUI.objects.all()
    serializer_class = serializers.CustomRuleTemplateUISerializer
    lookup_field = "pk"

    @action(detail=False, methods=["get"], url_path=r"schema/(?P<rule_name>[^/.]+)")
    def retrieve_by_rule(self, request, rule_name=None):
        """
        GET /api/rules/schema/<rule_name>/
        returns the schema JSON for the given rule_name
        """
        obj = get_object_or_404(models.CustomRuleTemplateUI, rule_name=rule_name)
        serializer = self.get_serializer(obj)
        return Response(serializer.data, status=status.HTTP_200_OK)


class AllRulesView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        context = {
            "success": 1,
            "message": "Rule names fetched successfully",
            "data": []
        }
        try:
            rules = models.CustomRuleTemplateUI.objects.all()
            rule_names = rules.values_list("rule_name", flat=True)  # gives QuerySet
            context["data"] = list(rule_names)  # convert to list for JSON response
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context)

class RuleAppliedView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    # CREATE (POST)
    def post(self, request):
        context = {
            "success": 1,
            "message": "Rule saved successfully",
            "data": {}
        }
        try:
            data = request.data
            # 1️⃣ Extract source_fields from payload
            source_fields = data.get("rule_applied_data", {}).get("source", {})
            spec_id = models.DataObject.objects.get(objectName=source_fields.get("spec")).id

            # 2️⃣ Find matching Specs record
            spec_obj = models.Specs.objects.filter(
                objectName=spec_id,
                tab=source_fields.get("tab"),
                field_id=source_fields.get("field"),
            ).first()

            if not spec_obj:
                context["success"] = 0
                context["message"] = "Spec not found for given source_fields"
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

            # 3️⃣ Inject spec_field_id before serializer validation
            data["spec"] = spec_obj.id

            serializer = serializers.RuleAppliedSerializer(
                data=data,
                context={"request": request}
            )
            if serializer.is_valid():
                serializer.save()
                context["data"] = serializer.data
            else:
                context["success"] = 0
                context["message"] = "Validation failed"
                context["errors"] = serializer.errors
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(context, status=status.HTTP_201_CREATED)

    # READ (GET Single Rule by ID)
    def get(self, request, pk=None):
        try:
            rule = models.RuleApplied.objects.select_related("spec").get(id=pk)
            serializer = serializers.RuleAppliedSerializer(rule)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except models.RuleApplied.DoesNotExist:
            return Response(
                {"success": 0, "message": "Rule not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    # UPDATE (PUT/PATCH)
    def put(self, request, pk=None):
        context = {"success": 1, "message": "Rule updated successfully", "data": {}}
        try:
            rule = models.RuleApplied.objects.get(id=pk)
            serializer = serializers.RuleAppliedSerializer(
                rule, data=request.data, partial=True, context={"request": request}
            )
            if serializer.is_valid():
                serializer.save()
                context["data"] = serializer.data
            else:
                context["success"] = 0
                context["message"] = "Validation failed"
                context["errors"] = serializer.errors
                return Response(context, status=status.HTTP_400_BAD_REQUEST)

        except models.RuleApplied.DoesNotExist:
            return Response(
                {"success": 0, "message": "Rule not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(context, status=status.HTTP_200_OK)

    # DELETE
    def delete(self, request, pk=None):
        try:
            rule = models.RuleApplied.objects.get(id=pk)
            rule.delete()
            return Response(
                {"success": 1, "message": "Rule deleted successfully"},
                status=status.HTTP_200_OK,
            )
        except models.RuleApplied.DoesNotExist:
            return Response(
                {"success": 0, "message": "Rule not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"success": 0, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class RuleAppliedListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        try:
            queryset = models.RuleApplied.objects.select_related("spec").all()

            # --- Filtering ---
            object_name = request.GET.get("objectName")
            if object_name:
                queryset = queryset.filter(spec__objectName__objectName=object_name)

            rule_applied = request.GET.get("rule_applied")
            if rule_applied:
                queryset = queryset.filter(rule_applied__icontains=rule_applied)

            created_after = request.GET.get("created_after")
            if created_after:
                queryset = queryset.filter(created_at__gte=created_after)

            created_before = request.GET.get("created_before")
            if created_before:
                queryset = queryset.filter(created_at__lte=created_before)

            # --- Search ---
            search = request.GET.get("search")
            if search:
                queryset = queryset.filter(
                    Q(description__icontains=search)
                    | Q(spec__tab__icontains=search)
                    | Q(spec__field_id__icontains=search)
                )

            # --- Ordering ---
            ordering = request.GET.get("ordering")
            if ordering:
                queryset = queryset.order_by(ordering)
            else:
                queryset = queryset.order_by("-created_at")

            serializer = serializers.RuleAppliedTableSerializer(queryset, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class RuleAppliedBySpecView(APIView):
    def get(self, request, spec_id):
        context = {
            "success": 1,
            "message": "Rules fetched successfully",
            "data": []
        }
        try:
            # SQLite-safe: values() + distinct()
            queryset = (
                models.RuleApplied.objects.filter(spec_id=spec_id)
                .values("rule_applied")
                .distinct()
            )

            serializer = serializers.RuleAppliedNameSerializer(queryset, many=True)

            # Extract only values → ["rule1", "rule2", ...]
            context["data"] = [item["rule_applied"] for item in serializer.data]

        except Exception as e:
            context["success"] = 0
            context["message"] = f"Failed to fetch rules: {str(e)}"

        return Response(context, status=status.HTTP_200_OK)

import pandas as pd
import io
from django.core.files.uploadedfile import InMemoryUploadedFile

def clean_excel(uploaded_file):
    """Return a cleaned Excel file object without 'mapping' sheet."""
    xls = pd.ExcelFile(uploaded_file)

    # write cleaned file to memory (BytesIO)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name in xls.sheet_names:
            if sheet_name.lower() != "mapping":
                df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)  # reset pointer to start

    # Wrap as InMemoryUploadedFile so Django treats it like a real upload
    cleaned_file = InMemoryUploadedFile(
        file=output,
        field_name="file",
        name="cleaned_file.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        size=output.getbuffer().nbytes,
        charset=None,
    )

    return cleaned_file


def handle_validated_upload(uploaded_file, obj, object_name):
    # ✅ Ensure directory exists
    base_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower())
    archive_dir = os.path.join(base_dir, "archive")
    os.makedirs(base_dir, exist_ok=True)
    os.makedirs(archive_dir, exist_ok=True)

    print('<<<<<<<<<<<<<<------')
    # print(os.getcwd())
    print(os.listdir(base_dir))
    print('--------->>>>>>>>>>>>>')

    # Paths
    main_file_path = os.path.join(base_dir, f"{object_name}.xlsx")

    # 🔥 Check if a main file already exists
    if os.path.exists(main_file_path):
        # Get last version
        last_file = models.DataFile.objects.filter(data_object=obj).order_by("-version").first()
        next_version = last_file.version + 1 if last_file else 1

        # Move existing main file → archive
        archive_name = f"{object_name}_v{next_version}.xlsx"
        archive_path = os.path.join(archive_dir, archive_name)
        shutil.move(main_file_path, archive_path)

        # Update previous DB record version
        models.DataFile.objects.filter(
            data_object=obj, file_name=f"{object_name}.xlsx"
        ).update(file_name=archive_name, version=next_version)

    # ✅ Save the new file as main (objectName.xlsx)
    with open(main_file_path, "wb+") as dest:
        for chunk in uploaded_file.chunks():
            dest.write(chunk)

    # 🔥 Create DB record for the new main file (no version)
    models.DataFile.objects.create(
        data_object=obj,
        file_name=f"{object_name}.xlsx",
        status=constants.STATUS_UPLOAD_SUCCESS,
        version=0,  # version 0 means "main/latest"
    )

    return Response(
        {
            "success": 1,
            "message": f"File uploaded and validated successfully for {object_name}",
            "file_name": f"{object_name}.xlsx",
            "status": constants.STATUS_UPLOAD_SUCCESS,
        },
        status=status.HTTP_200_OK,
    )

import tempfile
from django.core.files import File
import uuid, threading, time
from django.core.files.base import File
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import pandas as pd, tempfile
from .models import DataObject, Specs

# In-memory progress store (better: Redis/DB in production)
UPLOAD_PROGRESS = {}
class FileUploadAPIView(APIView):

    def get(self, request, id=None):
        """
        If object_id is given → return files for that DataObject.
        Otherwise → return all uploaded files.
        """
        context = {
            "success": 1,
            "message": "",
            "data": [],
        }

        try:
            object_id = id
            if object_id:
                files = DataFile.objects.filter(data_object_id=object_id).order_by("-uploaded_at")
                context["message"] = f"Files for DataObject ID {object_id} retrieved successfully"
            else:
                files = DataFile.objects.filter(version=0).order_by("-uploaded_at")
                context["message"] = "All files retrieved successfully"

            print('files...->>>>>>>>>>>>>>>>>>>>\n',files)
            serializer = serializers.DataFileSerializer(files, many=True)
            context["data"] = serializer.data

            return Response(context, status=status.HTTP_200_OK)

        except Exception as e:
            context["success"] = 0
            context["message"] = f"Error: {str(e)}"
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def post(self, request):
        # print("📥 Received POST request to handle-file API")
        serializer = FileUploadSerializer(data=request.data)
        if not serializer.is_valid():
            # print("❌ Serializer validation failed:", serializer.errors)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        object_name = serializer.validated_data["objectName"]
        uploaded_file = serializer.validated_data["file"]
        # print(f"✅ Serializer validated | objectName={object_name}, file={uploaded_file.name}")

        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_file_path = tmp.name
            for chunk in uploaded_file.chunks():
                tmp.write(chunk)
        # print(f"📂 Temp file saved at {temp_file_path}")

        # Validate objectName exists
        try:
            obj = models.DataObject.objects.get(objectName=object_name)
            # print(f"✅ DataObject found: {obj}")
        except models.DataObject.DoesNotExist:
            # print(f"❌ DataObject '{object_name}' not found in DB")
            return Response(
                {"success": 0, "message": f"Object '{object_name}' not found"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Specs check
        specs = Specs.objects.filter(objectName=obj)
        if not specs.exists():
            # print(f"❌ No specs defined for '{object_name}'")
            return Response(
                {"success": 0, "message": f"No specs defined for '{object_name}'"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        # print(f"✅ Specs found for '{object_name}' | count={specs.count()}")

        # Create unique upload ID
        upload_id = str(uuid.uuid4())
        UPLOAD_PROGRESS[upload_id] = {"progress": 0, "message": "Started processing...", "success": None}
        # print(f"🚀 Upload started | upload_id={upload_id}")

        # Process file in background
        def process_file():
            try:
                # Step 1: Building expected tabs
                # print(f"[{upload_id}] 🔄 Step 1: Building expected tabs")
                for p in range(1, 11):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Validating specs...", "success": None}
                    time.sleep(0.05)

                expected_tabs = {}
                for s in specs:
                    expected_tabs.setdefault(s.tab.lower(), set()).add(s.field_id.upper())
                # print(f"[{upload_id}] ✅ Expected tabs: {list(expected_tabs.keys())}")

                # Step 2: Reading Excel
                # print(f"[{upload_id}] 🔄 Reading Excel file {temp_file_path}")
                try:
                    xls = pd.ExcelFile(temp_file_path)
                    # print(f"[{upload_id}] ✅ Sheets found: {xls.sheet_names}")
                except Exception as e:
                    # print(f"[{upload_id}] ❌ Failed to read Excel file:", e)
                    UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": "Failed to read Excel file", "success": 0}
                    return

                for p in range(11, 31):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Preparing file...", "success": None}
                    time.sleep(0.05)

                # Step 3: Normalizing
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                    cleaned_file_path = tmp.name
                    with pd.ExcelWriter(cleaned_file_path, engine="openpyxl") as writer:
                        for sheet_name in xls.sheet_names:
                            if sheet_name.lower() == "mapping":
                                print(f"[{upload_id}] ⚠️ Skipping 'mapping' sheet")
                                continue
                            df = pd.read_excel(xls, sheet_name=sheet_name)
                            df.columns = [str(col).upper() for col in df.columns]
                            df.to_excel(writer, sheet_name=sheet_name.lower(), index=False)
                    # print(f"[{upload_id}] ✅ Normalized file written to {cleaned_file_path}")

                for p in range(31, 61):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Validating tabs & fields...", "success": None}
                    time.sleep(0.05)

                # Step 4: Validate tabs
                cleaned_xls = pd.ExcelFile(cleaned_file_path)
                file_tabs = set(cleaned_xls.sheet_names)
                # print(f"[{upload_id}] ✅ Normalized file sheets: {file_tabs}")

                expected_tab_names = set(expected_tabs.keys())
                missing_tabs = expected_tab_names - file_tabs
                if missing_tabs:
                    # print(f"[{upload_id}] ❌ Missing tabs: {missing_tabs}")
                    UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": f"Missing required tabs...❌", "success": 0}
                    return

                for tab, expected_fields in expected_tabs.items():
                    df = pd.read_excel(cleaned_xls, sheet_name=tab, nrows=1)
                    file_fields = set(df.columns.astype(str).str.upper())
                    missing_fields = expected_fields - file_fields
                    if missing_fields:
                        # print(f"[{upload_id}] ❌ Missing fields in {tab}: {missing_fields}")
                        UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": f"Missing fields in '{tab}'", "success": 0}
                        return
                # print(f"[{upload_id}] ✅ All required fields present")

                for p in range(61, 91):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Saving validated file...", "success": None}
                    time.sleep(0.05)

                # Step 5: Save normalized file
                with open(cleaned_file_path, "rb") as f:
                    cleaned_file = File(f, name=f"{object_name.lower()}_cleaned.xlsx")
                    handle_validated_upload(cleaned_file, obj, object_name)
                # print(f"[{upload_id}] 📤 File uploaded successfully to DB/storage")

                # Step 6: Finalizing
                for p in range(91, 101):
                    UPLOAD_PROGRESS[upload_id] = {"progress": p, "message": "Finalizing...", "success": None}
                    time.sleep(0.05)

                UPLOAD_PROGRESS[upload_id] = {"progress": 100, "message": "Upload completed successfully! 🎉", "success": 1}
                # print(f"[{upload_id}] 🎉 Upload finished successfully")

            except Exception as e:
                # print(f"[{upload_id}] ❌ Exception during processing:", e)
                UPLOAD_PROGRESS[upload_id] = {"progress": -1, "message": str(e), "success": 0}

        threading.Thread(target=process_file).start()
        return Response({"success": 1, "upload_id": upload_id}, status=status.HTTP_200_OK)

    def delete(self, request, id):
        try:
            file_obj = DataFile.objects.get(pk=id)
            temp = ''
            for i in str(file_obj.file_name):
                if i!='.':
                    temp = temp + i
                else:
                    break
            file_name = str(file_obj.file_name)   # e.g., "employees.xlsx"
            media_root = settings.MEDIA_ROOT + str(temp)     # e.g., "FileData/customer"
            file_path = os.path.join(media_root, file_name)  # Full path with filename

            archive_dir = os.path.join(media_root, "archive")
            deleted_dir = os.path.join(archive_dir, "delete")

            os.makedirs(archive_dir, exist_ok=True)
            os.makedirs(deleted_dir, exist_ok=True)

            # 🔹 Find last version
            last_version = (
                DataFile.objects.filter(data_object=file_obj.data_object)
                .order_by("-version")
                .first()
            )
            next_version = (last_version.version if last_version else 0) + 1

            base_name, ext = os.path.splitext(file_name)
            versioned_filename = f"{base_name}_v{next_version}{ext}"

            archived_path = os.path.join(archive_dir, versioned_filename)
            deleted_path = os.path.join(deleted_dir, versioned_filename)

            if os.path.exists(file_path):
                shutil.move(file_path, archived_path)   # Move file to archive
                shutil.copy2(archived_path, deleted_path)  # Copy also into delete

            # 🔹 Update DB
            file_obj.version = next_version
            file_obj.file_name = versioned_filename
            file_obj.save()

            models.DeletedFileRecord.objects.create(
                data_file=file_obj).save()


            serializer = serializers.DataFileSerializer(file_obj)
            return Response(
                {
                    "success": 1,
                    "message": "File archived and copied to deleted folder with new version",
                    "data": serializer.data,
                },
                status=status.HTTP_200_OK,
            )

        except DataFile.DoesNotExist:
            return Response(
                {"success": 0, "message": "File not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except Exception as e:
            return Response(
                {"success": 0, "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class UploadStatusView(APIView):
    def get(self, request, upload_id):
        progress = UPLOAD_PROGRESS.get(upload_id, 0)
        print(progress,'.....................................................................................')
        return Response({"upload_id": upload_id, "progress": progress})



from math import isfinite

def sanitize_data(data):
    if isinstance(data, dict):
        return {k: sanitize_data(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_data(i) for i in data]
    elif isinstance(data, float):
        return data if isfinite(data) else 0  # or None
    else:
        return data

# In your view



class DataFileLatestView(APIView):
    def get(self, request, object_id):
        try:
            mode = request.query_params.get("mode", "view")  # default = view
            print("\n[DEBUG] Incoming GET request ----------------------")
            print("[DEBUG] object_id:", object_id)
            print("[DEBUG] mode:", mode)
            if mode == 'download':
                latest_file = (
                models.DataFile.objects.filter(id=object_id, version=0)
                .first()
                )
            else:
                latest_file = (
                    models.DataFile.objects.filter(data_object_id=object_id, version=0)
                    .first()
                )

            print("[DEBUG] latest_file object:", latest_file)

            if not latest_file:
                print("[DEBUG] No DataFile found for object_id:", object_id)
                return Response(
                    {"success": 0, "message": "No file found", "data": {}},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Serialize the latest file
            serializer = serializers.DataFileSerializer(latest_file)
            safe_data = sanitize_data(serializer.data)
            print("[DEBUG] Serialized DataFile:", safe_data)

            # Extract file path
            file_name = str(latest_file.file_name)
            print("[DEBUG] file_name:", file_name)

            object_name = file_name.split(".")[0]
            print("[DEBUG] object_name (from file_name):", object_name)

            base_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower())
            print("[DEBUG] base_dir:", base_dir)

            file_path = os.path.normpath(os.path.join(base_dir, file_name))
            print("[DEBUG] file_path:", file_path)

            # ----------------
            # Mode 1: VIEW
            # ----------------
            if mode == "view":
                print("[DEBUG] Entered VIEW mode")
                data = {}
                if file_path.endswith(".xlsx") and os.path.exists(file_path):
                    print("[DEBUG] File exists, opening with pandas:", file_path)
                    with pd.ExcelFile(file_path) as xls:
                        print("[DEBUG] Sheets found:", xls.sheet_names)
                        for sheet in xls.sheet_names:
                            df = xls.parse(sheet)
                            fields = df.columns.tolist()
                            top_rows = df.head(30).to_dict(orient="records")
                            safe_rows = sanitize_data(top_rows)
                            data[sheet] = {
                                "fields": fields,
                                "rows": safe_rows
                            }
                            print(f"[DEBUG] Processed sheet: {sheet}, rows: {len(df)}")

                else:
                    print("[DEBUG] File missing or not .xlsx:", file_path)

                return Response(
                    {
                        "success": 1,
                        "message": "Latest file data retrieved successfully",
                        "data": {
                            "file_info": safe_data,
                            "excel_preview": data
                        },
                    },
                    status=status.HTTP_200_OK,
                )

            # ----------------
            # Mode 2: DOWNLOAD
            # ----------------
            elif mode == "download":
                print("[DEBUG] Entered DOWNLOAD mode")
                print("[DEBUG] Checking file path:", file_path)

                if not os.path.exists(file_path):
                    print("[DEBUG] File not found at path:", file_path)
                    return Response(
                        {"success": 0, "message": "File not found", "data": {}},
                        status=status.HTTP_404_NOT_FOUND,
                    )

                print("[DEBUG] File exists, preparing Excel for download...")
                xls = pd.ExcelFile(file_path)
                print("[DEBUG] Sheets found:", xls.sheet_names)

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    for sheet in xls.sheet_names:
                        df = xls.parse(sheet)
                        df.to_excel(writer, sheet_name=sheet, index=False)
                        print(f"[DEBUG] Copied sheet: {sheet}, rows: {len(df)}")

                    # Add mapping tab
                    print("[DEBUG] Fetching mapping specs for object_name:", object_name)
                    objectNameObj = models.DataObject.objects.filter(objectName=object_name).first()
                    print("[DEBUG] DataObject fetched:", objectNameObj)

                    if objectNameObj:
                        objectNameId = objectNameObj.id
                        spec_data = models.Specs.objects.filter(objectName=objectNameId).values()
                        spec_df = pd.DataFrame(spec_data)
                        print("[DEBUG] Spec rows:", len(spec_df))

                        spec_df = spec_df.drop(columns=["id", "position"], errors="ignore")
                        spec_df = spec_df.rename(columns={"objectName_id": "objectName"})
                        spec_df["objectName"] = object_name
                        spec_df.to_excel(writer, sheet_name="mapping", index=False)
                        print("[DEBUG] Added mapping tab")

                output.seek(0)
                print("[DEBUG] Returning Excel file as response")
                response = HttpResponse(
                    output,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                response["Content-Disposition"] = f'attachment; filename="{file_name}"'
                return response

            else:
                print("[DEBUG] Invalid mode:", mode)
                return Response(
                    {"success": 0, "message": "Invalid mode", "data": {}},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:
            print("[ERROR] Exception in get():", str(e))
            import traceback; traceback.print_exc()
            return Response(
                {"success": 0, "message": f"Error: {str(e)}", "data": {}},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

# views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Max
from .models import DataFile
from .serializers import DataFileSerializer
from django.db.models import Max, Q
from django.db.models import Exists, OuterRef
from django.db.models import Exists, OuterRef, Max, Q
class LatestValidatedFilesView(APIView):
    ''' Fetch latest validated (validation=1) file per DataObject '''
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    from django.db.models import Exists, OuterRef, Max

    def get(self, request):
        context = {
            "success": 0,
            "message": "Something went wrong.",
            "data": [],
        }
        try:
            # Step 1: Subquery to check if the data_object has any previous validation
            validated_subquery = models.DataFile.objects.filter(
                data_object=OuterRef('data_object'),
                validation=1
            )

            # Step 2: Fetch version=0 files with previous validation
            version0_files = models.DataFile.objects.filter(
                version=0
            ).annotate(
                has_previous_validation=Exists(validated_subquery)
            ).filter(has_previous_validation=True)

            # Step 3: Prepare a dict to track latest file per data_object
            latest_files_dict = {}

            # Add version=0 files first
            for file in version0_files:
                latest_files_dict[file.data_object] = file

            # Step 4: For data_objects without version=0 validated file, fetch latest validated file
            data_object_ids_with_v0 = version0_files.values_list('data_object', flat=True)
            fallback_files = (
                models.DataFile.objects
                .filter(validation=1)
                .exclude(data_object__in=data_object_ids_with_v0)
                .order_by('data_object', '-validated_at')
            )

            # Add fallback latest file per data_object
            for file in fallback_files:
                if file.data_object not in latest_files_dict:
                    latest_files_dict[file.data_object] = file

            # Step 5: Final list of files (unique per data_object)
            combined_files = list(latest_files_dict.values())

            serializer = serializers.DataFileSerializer(combined_files, many=True)

            if not serializer.data:
                context["message"] = "No validated files found."
                return Response(context, status=status.HTTP_404_NOT_FOUND)

            context["success"] = 1
            context["message"] = "Latest validated files fetched successfully."
            context["data"] = serializer.data

            return Response(context, status=status.HTTP_200_OK)

        except Exception as e:
            context["message"] = str(e)
            return Response(context, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# def creat_and_getworking_file_path(object_id):
#     latest_file = (
#                 models.DataFile.objects.filter(data_object_id=object_id, version=0)
#                 .first()
#             )
#     if not latest_file:
#         return 0

#     # Extract file path
#     file_name = str(latest_file.file_name)
#     object_name = file_name.split(".")[0]
#     base_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower())
#     file_path = os.path.normpath(os.path.join(base_dir, file_name))
#     return file_path

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from . import models
from . CustomValidationFiles.common_rules_validators import run_default_validators
from . CustomValidationFiles.custom_rule_validator import run_custom_rule_validation
from . file_utils import get_file_path_with_object_name
import time
from . working_file_manager import create_and_get_working_file_path,delete_working_directory
from django.utils import timezone

from . file_utils import get_target_specs
# #########################################################################
import uuid
from datetime import datetime
import threading

# Global in-memory store for progress objects
PROGRESS_STORE = {}

def create_progress(task_name: str):
    task_id = str(uuid.uuid4())
    progress_obj = {
        "id": task_id,
        "task_name": task_name,
        "progress": 0,
        "message": "Started",
        "created_at": datetime.now().isoformat(),
        "updated_at": datetime.now().isoformat()
    }
    PROGRESS_STORE[task_id] = progress_obj
    print(f"[PROGRESS] Created tracker: {task_id} for task: {task_name}")
    return task_id, progress_obj

def update_progress(task_id: str, progress: int, message: str = ""):
    if task_id in PROGRESS_STORE:
        PROGRESS_STORE[task_id]["progress"] = progress
        if message:
            PROGRESS_STORE[task_id]["message"] = message
        PROGRESS_STORE[task_id]["updated_at"] = datetime.now().isoformat()
        print(f"[PROGRESS] {task_id} -> {progress}% | {message}")

def get_progress(task_id: str):
    return PROGRESS_STORE.get(task_id, {"progress": 0, "message": ""})


def run_validation_in_background(task_id, data_object_id, request_data):
    """All heavy validation logic moved here."""
    try:
        update_progress(task_id, 5, "Fetching DataObject")
        print(f"[Thread] Fetching DataObject {data_object_id}")
        data_object = models.DataObject.objects.filter(id=data_object_id).first()
        if not data_object:
            update_progress(task_id, 6, "DataObject not found")
            print("[Thread ERROR] DataObject not found")
            return

        # Check own file
        update_progress(task_id, 6, "Checking main data file")
        print("[Thread] Checking main data file")
        own_file_exists = models.DataFile.objects.filter(data_object=data_object, version=0).exists()
        if not own_file_exists:
            update_progress(task_id, 7, f"Data file for '{data_object.objectName}' not found")
            print(f"[Thread ERROR] Data file for '{data_object.objectName}' not found")
            return

        # Check dependencies
        update_progress(task_id, 7, "Checking dependencies")
        print("[Thread] Checking dependencies")
        dependencies = data_object.dependencies or []
        rules_applied_qs = models.RuleApplied.objects.filter(spec__objectName=data_object.id)
        target_objects_dependencies = []
        for rule in rules_applied_qs:
            targets = get_target_specs(rule.rule_applied_data)
            target_objects_dependencies.extend(targets)
        dependencies = list(set(dependencies + target_objects_dependencies))
        update_progress(task_id, 8, "verifying depenedency files..")

        missing_files = []
        for dep_name in dependencies:
            dep_object = models.DataObject.objects.filter(objectName=dep_name).first()
            if not dep_object or not models.DataFile.objects.filter(data_object=dep_object, version=0).exists():
                missing_files.append(dep_name)
                print(f"[Thread] Missing dependency: {dep_name}")
        if missing_files:
            update_progress(task_id, 9, f"Missing dependencies: {', '.join(missing_files)}")
            print(f"[Thread ERROR] Missing dependencies: {', '.join(missing_files)}")
            return

        update_progress(task_id, 9, "Dependencies validated")
        print("[Thread] Dependencies validated")

        # Start default validations
        update_progress(task_id, 10, "Running default validations")
        paths = create_and_get_working_file_path(request_data.get("dataObjectId"))
        print(f"[Thread] Working paths: {paths}")
        resultLog1 = run_default_validators(
            file_path=paths.get('working_file_path'),
            log_file_path=paths.get('log_file_path'),
            primary_field=request_data.get("fieldId"),
            task_id=task_id,
            update_progress_fun=update_progress
        )
        update_progress(task_id, 65, "default Validation completed successfully")
        delete_working_directory(paths.get('working_file_path'))
        print("[Thread] Deleted working directory")

        # Prepare logs
        update_progress(task_id, 68, "Processing logs")
        source_file = get_file_path_with_object_name(data_object.objectName)
        log_file_path = paths.get('log_file_path')
        try:
            with open(log_file_path, "rb") as f:
                existing_log = pd.read_excel(f)
            print(f"[Thread] Existing log loaded. Rows: {len(existing_log)}")
        except FileNotFoundError:
            existing_log = pd.DataFrame(columns=["primary_field", "rule_data", "time"])
            print("[Thread] No existing log found, created empty log")

        # Run custom rule validations
       
        ###############################################
        update_progress(task_id, 70, "Running custom rule validations")
        new_logs_list = []
        rules_applied_qs = models.RuleApplied.objects.filter(spec__objectName=data_object.id)
        targets_obj = {}
        total_rules = rules_applied_qs.count()

        for i, rule in enumerate(rules_applied_qs):
            targets = get_target_specs(json_spec=rule.rule_applied_data)
            for obj in targets:
                targets_obj[obj] = get_file_path_with_object_name(obj)
            print(f"[Thread] Running custom rule {rule.id} on targets: {targets_obj}")

            resultLog2 = run_custom_rule_validation(
                rule_name=rule.rule_applied,
                json_spec=rule.rule_applied_data,
                source_file=source_file,
                target_files=targets_obj,
                rule_description=rule.description or ""
            )
            print(f"[Thread] Custom rule {rule.id} produced {len(resultLog2)} logs")
            new_logs_list.append(pd.DataFrame(resultLog2, columns=["primary_field", "rule_data", "time"]))

            # Update progress for this rule
            if task_id and total_rules > 0:
                incremental_progress = int((i + 1) / total_rules * 25)  # 25% for this loop
                update_progress(task_id, 70 + incremental_progress, f"Running custom rule {i + 1}/{total_rules}")

        ############
        #######
        ##########
        if new_logs_list:
            all_new_logs = pd.concat(new_logs_list, ignore_index=True)
            final_log = pd.concat([existing_log, all_new_logs], ignore_index=True)
        else:
            final_log = existing_log

        with pd.ExcelWriter(log_file_path, engine="openpyxl", mode="w") as writer:
            final_log.to_excel(writer, index=False)

        update_progress(task_id, 98, "written in log file successfully !")
        print("[Thread] Final log written to Excel")

        # Update DataFile validation status
        data_file = models.DataFile.objects.filter(data_object=data_object, version=0).first()
        if data_file:
            data_file.validation = 1
            data_file.validated_at = timezone.now()
            data_file.save()
            print("[Thread] DataFile validation updated")

        update_progress(task_id, 100, "Validation completed successfully")
        print("[Thread] Validation completed successfully")

    except Exception as e:
        update_progress(task_id, 0, f"Error: {str(e)}")
        print(f"[Thread ERROR] {str(e)}")


class PreValidationCheckAndValidationView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, data_object_id):
        print(f"\n[API] Validation API called with data_object_id: {data_object_id}")
        # Create progress tracker
        task_id, tracker = create_progress(f"Validation-{data_object_id}")

        # Start background thread
        thread = threading.Thread(
            target=run_validation_in_background,
            args=(task_id, data_object_id, request.data),
            daemon=True
        )
        thread.start()

        # Return immediately with task_id
        return Response({
            "success": 1,
            "message": "Validation started",
            "data": {"task_id": task_id}
        }, status=status.HTTP_200_OK)


class ValidationProgressView(APIView):
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, task_id):
        progress = PROGRESS_STORE.get(task_id, {"progress": 0, "message": "Task not started"})
        return Response({"success": 1, "data": progress})


import os
import glob
import pandas as pd
from django.conf import settings
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status


import os
import glob
import pandas as pd
from django.http import FileResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.conf import settings

class GetLatestLogDataView(APIView):
    """
    API to fetch top 30 rows of the latest log data for a given objectName.
    If `?download=1` is passed, returns the file for download.
    """
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, object_name):
        context = {
            "success": 1,
            "message": "Latest log data retrieved successfully",
            "data": []
        }
        try:
            # Log directory path
            log_dir = os.path.join(settings.MEDIA_ROOT, str(object_name).lower(), "Log")
            if not os.path.exists(log_dir):
                context["success"] = 0
                context["message"] = f"No log directory found for {object_name}"
                return Response(context, status=status.HTTP_404_NOT_FOUND)

            # Get all log files
            log_files = glob.glob(os.path.join(log_dir, "*.xlsx"))
            if not log_files:
                context["success"] = 0
                context["message"] = f"No log files found for {object_name}"
                return Response(context, status=status.HTTP_404_NOT_FOUND)

            # Pick latest log file
            latest_file = max(log_files, key=os.path.getmtime)

            # ✅ If download flag is passed → return file
            if request.query_params.get("download") == "1":
                return FileResponse(
                    open(latest_file, "rb"),
                    as_attachment=True,
                    filename=os.path.basename(latest_file),
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

            # Otherwise → return JSON response
            with open(latest_file, "rb") as f:
                df = pd.read_excel(f)

            df = df.fillna("").head(50)

            # context["file_name"] = os.path.basename(latest_file)
            # context["total_rows"] = len(df)
            context["data"] = df.to_dict(orient="records")

        except Exception as e:
            context["success"] = 0
            context["message"] = str(e)

        return Response(context, status=status.HTTP_200_OK if context["success"] else status.HTTP_500_INTERNAL_SERVER_ERROR)


import uuid, time, threading
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

UPLOAD_PROGRESS = {}  # store progress in memory (use Redis/DB in production)


def process_file(upload_id, file):
    """
    Simulated file processing in background thread.
    Replace with your actual logic (save file, parse, validate, etc.).
    """
    for i in range(0, 101, 20):  # fake steps
        time.sleep(1)
        UPLOAD_PROGRESS[upload_id] = i


@csrf_exempt
def handle_file(request):
    if request.method == "POST":
        file = request.FILES.get("file")
        if not file:
            return JsonResponse({"success": False, "message": "No file uploaded"})

        upload_id = str(uuid.uuid4())
        UPLOAD_PROGRESS[upload_id] = 0

        # run processing in background thread
        threading.Thread(target=process_file, args=(upload_id, file)).start()

        return JsonResponse({"success": True, "upload_id": upload_id})
    return JsonResponse({"success": False, "message": "Invalid method"})


def upload_status(request, upload_id):
    progress = UPLOAD_PROGRESS.get(upload_id, 0)
    return JsonResponse({"upload_id": upload_id, "progress": progress})


